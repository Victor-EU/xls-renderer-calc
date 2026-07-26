#!/usr/bin/env python3
"""Sample workbooks for the demo app.

Two of them are copied from the original spike (a three-sheet financial model,
once with LibreOffice-cached values and once without) because that pair is the
project's oldest ground truth. Two are new:

  hardcoded-total.xlsx  a model whose stated total does not equal the sum of its
                        parts — the case the computed-versus-cached diff exists
                        to catch, and the one a reviewer of LLM-written models
                        most wants flagged.
  formula-tour.xlsx     a wide sample of the function vocabulary a generated
                        model actually reaches for, with no cached values at all,
                        so the preview has to compute every number on screen.
"""

from __future__ import annotations

import os
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[2]
PUBLIC = ROOT / "apps" / "demo" / "public"

# Two of the demo fixtures predate this project: they come from the renderer
# spike this one grew out of, and they are the oldest ground truth here, which is
# why they are copied rather than regenerated. They are also already committed
# under apps/demo/public, so this only matters if you are rebuilding them — point
# XLSCALC_SPIKE_DIR at the spike's `public/` if you have it, and if you do not,
# the copies in the repository are what you want anyway.
SPIKE = Path(os.environ.get("XLSCALC_SPIKE_DIR", ""))

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
MONEY = '#,##0.0;[Red](#,##0.0)'
PCT = '0.0%'
BOLD = Font(bold=True)
TOP = Border(top=Side(style="thin"))


def hardcoded_total() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"
    ws.column_dimensions["A"].width = 26
    for c in "BCDE":
        ws.column_dimensions[c].width = 13

    ws["A1"] = "Revenue by segment"
    ws["A1"].font = Font(bold=True, size=13)
    for i, q in enumerate(["Q1", "Q2", "Q3", "Q4"]):
        cell = ws.cell(3, 2 + i, q)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.cell(3, 1, "Segment").font = HDR
    ws.cell(3, 1).fill = HDR_FILL

    rows = [("Platform", 412.0, 448.5, 470.2, 501.9), ("Services", 188.4, 191.0, 205.7, 219.3)]
    for r, (label, *vals) in enumerate(rows, start=4):
        ws.cell(r, 1, label)
        for i, v in enumerate(vals):
            cell = ws.cell(r, 2 + i, v)
            cell.number_format = MONEY

    total_row = 4 + len(rows)
    ws.cell(total_row, 1, "Total revenue").font = BOLD
    for i in range(4):
        col = 2 + i
        cell = ws.cell(total_row, col)
        # Q1–Q3 are honest sums. Q4 is typed in, and 40.0 too high: 501.9 + 219.3
        # is 721.2, not 761.2. Nothing about the rendered sheet reveals that —
        # the number is plausible, in the right place, formatted like its
        # neighbours. Only comparing it against the shape of its own row does.
        letter = chr(ord("B") + i)
        cell.value = 761.2 if i == 3 else f"=SUM({letter}4:{letter}{total_row - 1})"
        cell.number_format = MONEY
        cell.font = BOLD
        cell.border = TOP

    ws.cell(total_row + 2, 1, "Q4 growth vs Q3")
    ws.cell(total_row + 2, 2, f"=E{total_row}/D{total_row}-1").number_format = PCT
    return wb


def formula_tour() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14

    assumptions = [
        ("Opening revenue", 1200.0, MONEY),
        ("Growth rate", 0.085, PCT),
        ("Gross margin", 0.62, PCT),
        ("Opex ratio", 0.34, PCT),
        ("Tax rate", 0.23, PCT),
        ("Discount rate", 0.11, PCT),
        ("Start date", 46023, "yyyy-mm-dd"),
    ]
    ws["A1"] = "Assumptions"
    ws["A1"].font = Font(bold=True, size=13)
    for r, (label, value, fmt) in enumerate(assumptions, start=3):
        ws.cell(r, 1, label)
        cell = ws.cell(r, 2, value)
        cell.number_format = fmt

    model = wb.create_sheet("Model")
    model.column_dimensions["A"].width = 26
    for c in "BCDEF":
        model.column_dimensions[c].width = 13
    model.freeze_panes = "B4"

    model["A1"] = "Five-year model"
    model["A1"].font = Font(bold=True, size=13)
    for i in range(5):
        cell = model.cell(3, 2 + i, f"=EOMONTH(Assumptions!$B$9,{12 * (i + 1)})")
        cell.number_format = "mmm-yyyy"
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    model.cell(3, 1, "Line").font = HDR
    model.cell(3, 1).fill = HDR_FILL

    def row(r: int, label: str, formula, fmt=MONEY, bold=False):
        model.cell(r, 1, label)
        for i in range(5):
            col = chr(ord("B") + i)
            cell = model.cell(r, 2 + i, formula(col, i))
            cell.number_format = fmt
            if bold:
                cell.font = BOLD
        if bold:
            model.cell(r, 1).font = BOLD

    row(4, "Revenue", lambda col, i: (
        "=Assumptions!$B$3*(1+Assumptions!$B$4)"
        if i == 0
        else f"={chr(ord(col) - 1)}4*(1+Assumptions!$B$4)"
    ))
    row(5, "Gross profit", lambda col, i: f"=ROUND({col}4*Assumptions!$B$5,1)")
    row(6, "Operating expenses", lambda col, i: f"=-ROUND({col}4*Assumptions!$B$6,1)")
    row(7, "EBIT", lambda col, i: f"=SUM({col}5:{col}6)", bold=True)
    row(8, "Tax", lambda col, i: f"=-MAX(0,{col}7*Assumptions!$B$7)")
    row(9, "Net income", lambda col, i: f"={col}7+{col}8", bold=True)
    row(10, "Margin", lambda col, i: f"=IF({col}4=0,\"\",{col}9/{col}4)", fmt=PCT)
    row(11, "Growth", lambda col, i: (
        "=\"\"" if i == 0 else f"={col}4/{chr(ord(col) - 1)}4-1"
    ), fmt=PCT)

    summary = wb.create_sheet("Summary")
    summary.column_dimensions["A"].width = 30
    summary.column_dimensions["B"].width = 18
    summary["A1"] = "Summary"
    summary["A1"].font = Font(bold=True, size=13)
    kpis = [
        ("Total revenue", "=SUM(Model!B4:F4)", MONEY),
        ("Total net income", "=SUM(Model!B9:F9)", MONEY),
        ("Average margin", "=AVERAGE(Model!B10:F10)", PCT),
        ("Best year", "=MAX(Model!B9:F9)", MONEY),
        ("NPV of net income", "=NPV(Assumptions!B8,Model!B9:F9)", MONEY),
        ("Years to double", "=PDURATION(Assumptions!B4,1,2)", "0.0"),
        ("Payback year", '=IFERROR(MATCH(TRUE,INDEX(Model!B9:F9>0,0),0),"never")', "0"),
        ("Revenue over 1,500", '=COUNTIF(Model!B4:F4,">1500")', "0"),
        ("Weighted revenue", "=SUMPRODUCT(Model!B4:F4,Model!B10:F10)", MONEY),
        ("Label", '="Model through "&TEXT(Model!F3,"yyyy")', "General"),
    ]
    for r, (label, formula, fmt) in enumerate(kpis, start=3):
        summary.cell(r, 1, label)
        cell = summary.cell(r, 2, formula)
        cell.number_format = fmt
    return wb


def main() -> None:
    PUBLIC.mkdir(parents=True, exist_ok=True)
    for name in ("financial-model.xlsx", "financial-model-nocache.xlsx"):
        src = SPIKE / name if SPIKE.name else None
        if src is not None and src.exists():
            shutil.copy(src, PUBLIC / name)
            print(f"  copied  {name}")
        elif (PUBLIC / name).exists():
            print(f"  kept    {name} (already in the repository)")
        else:
            print(f"  MISSING {name} — set XLSCALC_SPIKE_DIR to the spike's public/ dir")

    hardcoded_total().save(PUBLIC / "hardcoded-total.xlsx")
    print("  wrote   hardcoded-total.xlsx")
    formula_tour().save(PUBLIC / "formula-tour.xlsx")
    print("  wrote   formula-tour.xlsx")


if __name__ == "__main__":
    main()
