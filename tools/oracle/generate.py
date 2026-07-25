#!/usr/bin/env python3
"""Build the oracle fixtures.

    probe suites (suites.py)
        ├─► build/<suite>.xlsx        formulas + grid, no cached values
        │        └─► soffice --headless --convert-to xlsx  (recalculates)
        │                 └─► build/recalc/<suite>.xlsx
        │                          └─► build/expected.json   the oracle
        └─► build/spec.json           the same workbook, for the engine to build

LibreOffice is not Excel, and on financial functions and edge-case rounding they
diverge. That is a known limit of this oracle (design §12.5), handled two ways:
divergences are recorded in `divergences.json` as documented expectations rather
than silent failures, and the financial functions carry hand-checked unit tests
in addition to this harness.
"""

from __future__ import annotations

import datetime as dt
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from suites import BASE_GRID, SUITES

ROOT = Path(__file__).resolve().parent
BUILD = ROOT / "build"
RECALC = BUILD / "recalc"

# Probes start well clear of the shared grid so that a whole-column probe like
# `SUM(A:A)` cannot accidentally include the probe cells themselves — which
# would be a real circular reference, and a confusing way to discover it.
PROBE_COL = 30  # column AD
SHEET = "Probes"

# Functions added to Excel after 2007 are stored in OOXML under the `_xlfn.`
# namespace. openpyxl writes formula text verbatim, so without this every one of
# them reads back as `#NAME?` from LibreOffice — which looks exactly like an
# oracle divergence and is really a fixture bug.
XLFN = {
    "IFS", "SWITCH", "MAXIFS", "MINIFS", "TEXTJOIN", "CONCAT", "XOR", "IFNA",
    "XLOOKUP", "XMATCH", "DAYS", "RRI", "PDURATION", "UNICHAR", "UNICODE",
    "NUMBERVALUE", "CEILING.MATH", "FLOOR.MATH", "ISFORMULA",
    "STDEV.S", "STDEV.P", "VAR.S", "VAR.P", "RANK.EQ", "RANK.AVG",
    "PERCENTILE.INC", "PERCENTILE.EXC", "QUARTILE.INC", "QUARTILE.EXC",
}

_FN_RE = re.compile(r"(?<![A-Za-z0-9_.])([A-Z][A-Z0-9.]*)\s*\(")


def to_ooxml(formula: str) -> str:
    """Namespace post-2007 function names the way Excel stores them."""
    return _FN_RE.sub(lambda m: (f"_xlfn.{m.group(1)}(" if m.group(1) in XLFN else m.group(0)), formula)


SOFFICE_CANDIDATES = [
    "soffice",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
]


def find_soffice() -> str:
    for candidate in SOFFICE_CANDIDATES:
        resolved = shutil.which(candidate) or (candidate if Path(candidate).exists() else None)
        if resolved:
            return resolved
    sys.exit("LibreOffice not found — install it or set one of: " + ", ".join(SOFFICE_CANDIDATES))


def build_workbook(formulas: list[str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    for addr, value in BASE_GRID.items():
        if value is None:
            continue  # a blank cell must stay genuinely absent, not an empty string
        ws[addr] = value
    for i, formula in enumerate(formulas, start=1):
        ws.cell(row=i, column=PROBE_COL, value=to_ooxml(formula))
    return wb


def recalc(soffice: str, src: Path) -> Path:
    RECALC.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            soffice,
            "--headless",
            "--norestore",
            "--convert-to",
            "xlsx:Calc MS Excel 2007 XML",
            "--outdir",
            str(RECALC),
            str(src),
        ],
        check=True,
        capture_output=True,
        timeout=180,
    )
    out = RECALC / src.name
    if not out.exists():
        sys.exit(f"LibreOffice produced no output for {src.name}")
    return out


def read_expected(path: Path, count: int) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb[SHEET]
    out = []
    for i in range(1, count + 1):
        v = ws.cell(row=i, column=PROBE_COL).value
        out.append(encode(v))
    return out


# Excel serial 0 is 1899-12-30 because of the phantom 1900-02-29; dates before
# 1900-03-01 sit one day off that line. Both branches are needed to convert
# LibreOffice's date-formatted answers back into the serials the engine returns.
_EPOCH = dt.date(1899, 12, 30)
_LEAP_BUG_CUTOFF = dt.date(1900, 3, 1)


def date_to_serial(d: dt.date) -> int:
    return (d - _EPOCH).days if d >= _LEAP_BUG_CUTOFF else (d - dt.date(1899, 12, 31)).days


def encode(v) -> dict:
    """Tag the value so the Node side can tell blank from "" from an error."""
    if v is None:
        return {"t": "blank"}
    if isinstance(v, bool):
        return {"t": "bool", "v": v}
    if isinstance(v, (int, float)):
        return {"t": "num", "v": float(v)}
    # LibreOffice applies a date or time format to the results of DATE()/TIME(),
    # so openpyxl hands those back as datetime objects rather than serials.
    if isinstance(v, dt.datetime):
        frac = (v.hour * 3600 + v.minute * 60 + v.second) / 86400
        return {"t": "num", "v": date_to_serial(v.date()) + frac}
    if isinstance(v, dt.date):
        return {"t": "num", "v": float(date_to_serial(v))}
    if isinstance(v, dt.time):
        return {"t": "num", "v": (v.hour * 3600 + v.minute * 60 + v.second) / 86400}
    if isinstance(v, dt.timedelta):
        return {"t": "num", "v": v.total_seconds() / 86400}
    if isinstance(v, str):
        if v.startswith("#") and v.endswith(("!", "?", "A")):
            return {"t": "err", "v": v}
        return {"t": "str", "v": v}
    return {"t": "other", "v": str(v)}


def main() -> None:
    soffice = find_soffice()
    BUILD.mkdir(parents=True, exist_ok=True)

    spec = {"grid": {k: v for k, v in BASE_GRID.items()}, "probeColumn": PROBE_COL, "sheet": SHEET, "suites": {}}
    expected: dict[str, list[dict]] = {}

    for name, formulas in SUITES.items():
        src = BUILD / f"{name}.xlsx"
        build_workbook(formulas).save(src)
        out = recalc(soffice, src)
        expected[name] = read_expected(out, len(formulas))
        spec["suites"][name] = formulas
        computed = sum(1 for e in expected[name] if e["t"] != "blank")
        print(f"  {name:16s} {len(formulas):4d} probes  {computed:4d} computed by LibreOffice")

    (BUILD / "spec.json").write_text(json.dumps(spec, indent=1))
    (BUILD / "expected.json").write_text(json.dumps(expected, indent=1))
    total = sum(len(f) for f in SUITES.values())
    print(f"\n{total} probes across {len(SUITES)} suites → build/spec.json + build/expected.json")
    print(f"probe column: {get_column_letter(PROBE_COL)}")


if __name__ == "__main__":
    main()
