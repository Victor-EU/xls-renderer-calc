"""M04 — Linked three-statement model with a balance check.

Income statement, balance sheet and cash flow statement, wired together the
normal way: working capital comes off the balance sheet into the cash flow, cash
comes back off the cash flow onto the balance sheet, and retained earnings rolls
forward off the income statement.

What it stresses: dependency ordering across three sheets where the natural
reading order is not the evaluation order — the balance sheet's cash line is
computed from a cash flow statement that is itself computed from the balance
sheet's working capital. An engine that evaluates in file order gets this wrong
and produces a balance sheet that does not balance.

The check row is the real assertion. A three-statement model that balances by
accident is not a thing that happens; if `Checks` says OK for all five years,
the whole chain evaluated in the right order.
"""

from __future__ import annotations

from openpyxl import Workbook

from common import (
    BOLD, Grid, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY, PCT, SUB_FILL,
    TOTAL_BORDER, col, header_row, note, title, widths,
)

YEARS = 5
FIRST = 2


def drivers(wb: Workbook):
    ws = wb.active
    ws.title = "Drivers"
    widths(ws, a=36, b=14)
    title(ws, "Operating drivers and opening balances")

    rows = [
        ("Opening revenue", 2400.0, MONEY),
        ("Revenue growth", 0.090, PCT),
        ("Gross margin", 0.580, PCT),
        ("Operating expenses, % of revenue", 0.360, PCT),
        ("Depreciation, % of opening gross PP&E", 0.110, PCT),
        ("Capex, % of revenue", 0.055, PCT),
        ("Days sales outstanding", 52, INT),
        ("Days inventory outstanding", 38, INT),
        ("Days payable outstanding", 46, INT),
        ("Tax rate", 0.250, PCT),
        ("Interest rate on debt", 0.058, PCT),
        ("Interest earned on cash", 0.021, PCT),
        ("Dividend payout, % of net income", 0.200, PCT),
        ("Scheduled debt repayment, per year", 50.0, MONEY),
        ("Days in year", 365, INT),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value)
        c.number_format = fmt
        c.font = INPUT_FONT

    ws.cell(19, 1, "Opening balance sheet").font = BOLD
    ws.cell(19, 1).fill = SUB_FILL
    opening = [
        ("Cash", 180.0),
        ("Accounts receivable", 340.0),
        ("Inventory", 145.0),
        ("Gross property, plant & equipment", 1250.0),
        ("Accumulated depreciation", -420.0),
        ("Accounts payable", 165.0),
        ("Debt", 600.0),
        ("Share capital", 300.0),
    ]
    for i, (label, value) in enumerate(opening, start=20):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value)
        c.number_format = MONEY
        c.font = INPUT_FONT

    # Retained earnings is the opening plug — it is what makes the opening
    # balance sheet balance, and writing it as a formula means it stays right if
    # any other opening balance changes.
    ws.cell(28, 1, "Retained earnings (opening plug)").font = BOLD
    ws.cell(28, 2, "=(B20+B21+B22+B23+B24)-(B25+B26)-B27").number_format = MONEY
    ws.cell(29, 1, "Opening balance check").font = BOLD
    ws.cell(29, 2,
            '=IF(ABS((B20+B21+B22+B23+B24)-(B25+B26+B27+B28))<0.01,"balanced","ERROR")')
    return ws


def income_statement(wb: Workbook):
    ws = wb.create_sheet("IS")
    widths(ws, a=34)
    for i in range(YEARS):
        ws.column_dimensions[col(FIRST + i)].width = 13
    title(ws, "Income statement")
    ws.freeze_panes = "B5"
    header(ws)

    g = Grid(ws, FIRST, YEARS, start_row=6)
    prev = lambda i: col(FIRST + i - 1)  # noqa: E731

    g.line("revenue", "Revenue",
           lambda i, c: ("=ROUND(Drivers!$B$3*(1+Drivers!$B$4),1)" if i == 0
                         else f"=ROUND({prev(i)}{g.R['revenue']}*(1+Drivers!$B$4),1)"), bold=True)
    g.line("cogs", "Cost of goods sold",
           lambda i, c: f"=-ROUND({c}{g.R['revenue']}*(1-Drivers!$B$5),1)")
    g.line("gp", "Gross profit",
           lambda i, c: f"={c}{g.R['revenue']}+{c}{g.R['cogs']}", bold=True)
    g.line("gm", "Gross margin",
           lambda i, c: f'=IFERROR({c}{g.R["gp"]}/{c}{g.R["revenue"]},"")', PCT)
    g.line("opex", "Operating expenses",
           lambda i, c: f"=-ROUND({c}{g.R['revenue']}*Drivers!$B$6,1)")
    g.line("ebitda", "EBITDA",
           lambda i, c: f"={c}{g.R['gp']}+{c}{g.R['opex']}", bold=True)
    # Depreciation runs off the *opening* gross PP&E, which is last year's
    # closing balance. Charging it on the closing balance would make the model
    # circular, since closing PP&E depends on this year's capex.
    g.line("dep", "Depreciation",
           lambda i, c: ("=-ROUND(Drivers!$B$23*Drivers!$B$7,1)" if i == 0
                         else f"=-ROUND(BS!{prev(i)}$8*Drivers!$B$7,1)"))
    g.line("ebit", "EBIT", lambda i, c: f"={c}{g.R['ebitda']}+{c}{g.R['dep']}", bold=True)
    g.line("int_exp", "Interest expense",
           lambda i, c: ("=-ROUND(Drivers!$B$26*Drivers!$B$13,1)" if i == 0
                         else f"=-ROUND(BS!{prev(i)}$16*Drivers!$B$13,1)"))
    g.line("int_inc", "Interest income",
           lambda i, c: ("=ROUND(Drivers!$B$20*Drivers!$B$14,1)" if i == 0
                         else f"=ROUND(BS!{prev(i)}$6*Drivers!$B$14,1)"))
    g.line("ebt", "Profit before tax",
           lambda i, c: f"={c}{g.R['ebit']}+{c}{g.R['int_exp']}+{c}{g.R['int_inc']}")
    g.line("tax", "Income tax",
           lambda i, c: f"=-ROUND(MAX(0,{c}{g.R['ebt']})*Drivers!$B$12,1)")
    g.line("ni", "Net income", lambda i, c: f"={c}{g.R['ebt']}+{c}{g.R['tax']}", bold=True)
    g.gap()
    g.line("div", "Dividends declared",
           lambda i, c: f"=-ROUND(MAX(0,{c}{g.R['ni']})*Drivers!$B$15,1)")
    g.line("retained", "Retained this year",
           lambda i, c: f"={c}{g.R['ni']}+{c}{g.R['div']}", bold=True)
    g.line("netmargin", "Net margin",
           lambda i, c: f'=IFERROR({c}{g.R["ni"]}/{c}{g.R["revenue"]},"")', PCT)
    g.render()
    for i in range(YEARS):
        ws.cell(g.R["ni"], FIRST + i).border = TOTAL_BORDER
    return ws, g


def balance_sheet(wb: Workbook, isg: Grid, cfg_rows: dict):
    """Fixed row layout — the IS references BS rows 6, 8 and 16 by number, so
    those three lines are pinned rather than laid out by the Grid helper."""
    ws = wb.create_sheet("BS")
    widths(ws, a=34)
    for i in range(YEARS):
        ws.column_dimensions[col(FIRST + i)].width = 13
    title(ws, "Balance sheet")
    ws.freeze_panes = "B5"
    header(ws)

    prev = lambda i: col(FIRST + i - 1)  # noqa: E731
    R = {
        "cash": 6, "ar": 7, "gross_ppe": 8, "accdep": 9, "inventory": 10,
        "net_ppe": 11, "total_assets": 13, "ap": 15, "debt": 16,
        "total_liab": 17, "capital": 19, "retained": 20, "total_equity": 21,
        "total_le": 23, "check": 25, "checktext": 26,
    }
    labels = {
        6: "Cash", 7: "Accounts receivable", 8: "Gross PP&E",
        9: "Accumulated depreciation", 10: "Inventory", 11: "Net PP&E",
        13: "Total assets", 15: "Accounts payable", 16: "Debt",
        17: "Total liabilities", 19: "Share capital", 20: "Retained earnings",
        21: "Total equity", 23: "Total liabilities and equity",
        25: "Balance check (assets less L+E)", 26: "Status",
    }
    for r, label in labels.items():
        c = ws.cell(r, 1, label)
        if r in (11, 13, 17, 21, 23, 25, 26):
            c.font = BOLD
    ws.cell(5, 1, "Assets").font = BOLD
    ws.cell(5, 1).fill = SUB_FILL
    ws.cell(14, 1, "Liabilities and equity").font = BOLD
    ws.cell(14, 1).fill = SUB_FILL

    for i in range(YEARS):
        c = col(FIRST + i)
        f = {
            R["cash"]: f"=CF!{c}{cfg_rows['closing']}",
            R["ar"]: f"=ROUND(IS!{c}{isg.R['revenue']}*Drivers!$B$9/Drivers!$B$17,1)",
            R["inventory"]: f"=ROUND(-IS!{c}{isg.R['cogs']}*Drivers!$B$10/Drivers!$B$17,1)",
            R["gross_ppe"]: (f"=Drivers!$B$23-CF!{c}{cfg_rows['capex']}" if i == 0
                             else f"={prev(i)}{R['gross_ppe']}-CF!{c}{cfg_rows['capex']}"),
            R["accdep"]: (f"=Drivers!$B$24+IS!{c}{isg.R['dep']}" if i == 0
                          else f"={prev(i)}{R['accdep']}+IS!{c}{isg.R['dep']}"),
            R["net_ppe"]: f"={c}{R['gross_ppe']}+{c}{R['accdep']}",
            R["total_assets"]: (f"={c}{R['cash']}+{c}{R['ar']}+{c}{R['inventory']}"
                                f"+{c}{R['net_ppe']}"),
            R["ap"]: f"=ROUND(-IS!{c}{isg.R['cogs']}*Drivers!$B$11/Drivers!$B$17,1)",
            R["debt"]: (f"=MAX(0,Drivers!$B$26-Drivers!$B$16)" if i == 0
                        else f"=MAX(0,{prev(i)}{R['debt']}-Drivers!$B$16)"),
            R["total_liab"]: f"={c}{R['ap']}+{c}{R['debt']}",
            R["capital"]: "=Drivers!$B$27",
            R["retained"]: (f"=Drivers!$B$28+IS!{c}{isg.R['retained']}" if i == 0
                            else f"={prev(i)}{R['retained']}+IS!{c}{isg.R['retained']}"),
            R["total_equity"]: f"={c}{R['capital']}+{c}{R['retained']}",
            R["total_le"]: f"={c}{R['total_liab']}+{c}{R['total_equity']}",
            R["check"]: f"=ROUND({c}{R['total_assets']}-{c}{R['total_le']},4)",
        }
        for r, formula in f.items():
            cell = ws.cell(r, FIRST + i, formula)
            cell.number_format = MONEY
            if r in (11, 13, 17, 21, 23, 25):
                cell.font = BOLD
        ws.cell(R["checktext"], FIRST + i,
                f'=IF(ABS({c}{R["check"]})<0.05,"OK","OUT BY "&TEXT({c}{R["check"]},"0.0000"))')
        ws.cell(R["total_assets"], FIRST + i).border = TOTAL_BORDER
        ws.cell(R["total_le"], FIRST + i).border = TOTAL_BORDER
    return ws, R


def cash_flow(wb: Workbook, isg: Grid):
    ws = wb.create_sheet("CF")
    widths(ws, a=34)
    for i in range(YEARS):
        ws.column_dimensions[col(FIRST + i)].width = 13
    title(ws, "Cash flow statement")
    ws.freeze_panes = "B5"
    header(ws)

    prev = lambda i: col(FIRST + i - 1)  # noqa: E731
    BSR = {"cash": 6, "ar": 7, "ap": 15, "inventory": 10, "debt": 16}

    g = Grid(ws, FIRST, YEARS, start_row=6)
    g.line("ni", "Net income", lambda i, c: f"=IS!{c}{isg.R['ni']}", bold=True)
    g.line("dep", "Add back depreciation", lambda i, c: f"=-IS!{c}{isg.R['dep']}")
    g.line("dar", "(Increase) in receivables",
           lambda i, c: (f"=-(BS!{c}{BSR['ar']}-Drivers!$B$21)" if i == 0
                         else f"=-(BS!{c}{BSR['ar']}-BS!{prev(i)}{BSR['ar']})"))
    g.line("dinv", "(Increase) in inventory",
           lambda i, c: (f"=-(BS!{c}{BSR['inventory']}-Drivers!$B$22)" if i == 0
                         else f"=-(BS!{c}{BSR['inventory']}-BS!{prev(i)}{BSR['inventory']})"))
    g.line("dap", "Increase in payables",
           lambda i, c: (f"=BS!{c}{BSR['ap']}-Drivers!$B$25" if i == 0
                         else f"=BS!{c}{BSR['ap']}-BS!{prev(i)}{BSR['ap']}"))
    g.line("cfo", "Cash from operations",
           lambda i, c: f"=SUM({c}{g.R['ni']}:{c}{g.R['dap']})", bold=True)
    g.gap()
    g.line("capex", "Capital expenditure",
           lambda i, c: f"=-ROUND(IS!{c}{isg.R['revenue']}*Drivers!$B$8,1)")
    g.line("cfi", "Cash from investing", lambda i, c: f"={c}{g.R['capex']}", bold=True)
    g.gap()
    g.line("repay", "Debt repayment",
           lambda i, c: (f"=-(Drivers!$B$26-BS!{c}{BSR['debt']})" if i == 0
                         else f"=-(BS!{prev(i)}{BSR['debt']}-BS!{c}{BSR['debt']})"))
    g.line("div", "Dividends paid", lambda i, c: f"=IS!{c}{isg.R['div']}")
    g.line("cff", "Cash from financing",
           lambda i, c: f"={c}{g.R['repay']}+{c}{g.R['div']}", bold=True)
    g.gap()
    g.line("net", "Net change in cash",
           lambda i, c: f"={c}{g.R['cfo']}+{c}{g.R['cfi']}+{c}{g.R['cff']}", bold=True)
    g.line("opening", "Cash, opening",
           lambda i, c: ("=Drivers!$B$20" if i == 0 else f"={prev(i)}{g.R['closing']}"))
    g.line("closing", "Cash, closing",
           lambda i, c: f"={c}{g.R['opening']}+{c}{g.R['net']}", bold=True)
    g.render()
    for i in range(YEARS):
        ws.cell(g.R["closing"], FIRST + i).border = TOTAL_BORDER
    return ws, g


def checks(wb: Workbook, isg: Grid, bsr: dict, cfg: Grid):
    ws = wb.create_sheet("Checks")
    widths(ws, a=40)
    for i in range(YEARS):
        ws.column_dimensions[col(FIRST + i)].width = 15
    title(ws, "Integrity checks")
    note(ws, 2, "Every line must read OK. Anything else means the statements do not tie.")
    header(ws, row=4)

    def retained_roll(i: int, c: str) -> str:
        opening = ("Drivers!$B$28" if i == 0
                   else f"BS!{col(FIRST + i - 1)}{bsr['retained']}")
        drift = f"BS!{c}{bsr['retained']}-{opening}-IS!{c}{isg.R['retained']}"
        return f'=IF(ABS({drift})<0.01,"OK","MISMATCH")'

    def cash_ties(i: int, c: str) -> str:
        drift = f"BS!{c}{bsr['cash']}-CF!{c}{cfg.R['closing']}"
        return f'=IF(ABS({drift})<0.01,"OK","OUT BY "&TEXT({drift},"0.00"))'

    tests = [
        ("Balance sheet balances", lambda i, c: f"=BS!{c}{bsr['checktext']}"),
        ("Cash ties to cash flow statement", cash_ties),
        ("Retained earnings rolls forward", retained_roll),
        ("Net PP&E is not negative",
         lambda i, c: f'=IF(BS!{c}{bsr["net_ppe"]}>=0,"OK","NEGATIVE")'),
        ("Cash is not negative",
         lambda i, c: f'=IF(BS!{c}{bsr["cash"]}>=0,"OK","OVERDRAWN")'),
        ("Debt never goes below zero",
         lambda i, c: f'=IF(BS!{c}{bsr["debt"]}>=0,"OK","NEGATIVE")'),
        ("Net income is positive",
         lambda i, c: f'=IF(IS!{c}{isg.R["ni"]}>0,"OK","LOSS")'),
    ]
    for j, (label, fn) in enumerate(tests):
        r = 6 + j
        ws.cell(r, 1, label)
        for i in range(YEARS):
            ws.cell(r, FIRST + i, fn(i, col(FIRST + i)))

    last = 5 + len(tests)
    grid = f"B6:{col(FIRST + YEARS - 1)}{last}"
    ws.cell(last + 2, 1, "Checks passed").font = BOLD
    ws.cell(last + 2, 2, f'=COUNTIF({grid},"OK")').number_format = INT
    ws.cell(last + 3, 1, "Checks total").font = BOLD
    ws.cell(last + 3, 2, f"=COUNTA({grid})").number_format = INT
    ws.cell(last + 4, 1, "Model status").font = BOLD
    ws.cell(last + 4, 2,
            f'=IF(B{last + 2}=B{last + 3},"MODEL TIES",'
            f'"REVIEW — "&TEXT(B{last + 3}-B{last + 2},"0")&" checks failed")')
    ws.cell(last + 6, 1, "Largest balance-sheet discrepancy").font = BOLD
    ws.cell(last + 6, 2,
            f"=MAX(ABS(BS!B{bsr['check']}),ABS(BS!C{bsr['check']}),ABS(BS!D{bsr['check']}),"
            f"ABS(BS!E{bsr['check']}),ABS(BS!F{bsr['check']}))").number_format = "0.0000"
    return ws


def header(ws, row: int = 4):
    ws.cell(row, 1, "Fiscal year").font = HDR_FONT
    ws.cell(row, 1).fill = HDR_FILL
    for i in range(YEARS):
        c = ws.cell(row, FIRST + i, f'="FY"&(2026+{i})')
        c.font = HDR_FONT
        c.fill = HDR_FILL


def build() -> Workbook:
    wb = Workbook()
    drivers(wb)
    _, isg = income_statement(wb)
    # CF is laid out before BS so BS knows where the closing-cash line sits;
    # CF references BS by pinned row numbers, which are asserted below.
    ws_cf, cfg = cash_flow(wb, isg)
    _, bsr = balance_sheet(wb, isg, cfg.R)
    assert (bsr["cash"], bsr["ar"], bsr["ap"], bsr["inventory"], bsr["debt"],
            bsr["gross_ppe"]) == (6, 7, 15, 10, 16, 8), "BS row layout moved — CF and IS pin these"
    checks(wb, isg, bsr, cfg)
    wb.move_sheet("CF", offset=1)
    return wb
