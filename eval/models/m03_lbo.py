"""M03 — Leveraged buyout with a debt schedule and cash sweep.

Sources and uses with the sponsor's equity as the plug, a five-year operating
case, a four-tranche debt schedule with mandatory amortisation and a cash sweep
waterfall, covenant tests, and a returns attribution split between EBITDA
growth, multiple change and deleveraging.

What it stresses: MIN/MAX waterfalls (the sweep cascades revolver → TLA → TLB,
so each tranche's prepayment depends on what the ones above it absorbed), a
five-link year-over-year chain across two sheets, IRR both algebraically and via
the function, and INDEX with a computed column index in the returns matrix.

Interest is charged on the *beginning* balance. That is the convention a careful
modeller uses precisely because charging it on the average balance makes the
model circular — and the circular version is in M10, on purpose.
"""

from __future__ import annotations

from openpyxl import Workbook

from common import (
    BOLD, Grid, HDR_FILL, HDR_FONT, INPUT_FONT, MONEY, MULT, PCT, PCT2,
    SUB_FILL, TOTAL_BORDER, col, header_row, note, title, widths,
)

YEARS = 5
FIRST = 2


def assumptions(wb: Workbook):
    ws = wb.active
    ws.title = "Assumptions"
    widths(ws, a=36, b=14)
    title(ws, "LBO assumptions — Project Halyard")
    note(ws, 2, "USD millions. Turns are multiples of LTM EBITDA.")

    rows = [
        ("LTM EBITDA", 240.0, MONEY),
        ("Entry multiple", 9.5, MULT),
        ("Transaction fees, % of EV", 0.020, PCT),
        ("Financing fees, % of debt", 0.025, PCT),
        ("Existing net debt refinanced", 380.0, MONEY),
        ("Minimum cash balance", 10.0, MONEY),
        ("Cash sweep, % of excess", 0.75, PCT),
        ("Revolver capacity", 60.0, MONEY),
        ("Base rate (SOFR)", 0.0435, PCT2),
        ("Revolver spread", 0.0300, PCT2),
        ("Revolver commitment fee", 0.0050, PCT2),
        ("Term Loan A, turns", 2.0, MULT),
        ("Term Loan A, spread", 0.0325, PCT2),
        ("Term Loan A, annual amortisation", 0.10, PCT),
        ("Term Loan B, turns", 2.5, MULT),
        ("Term Loan B, spread", 0.0425, PCT2),
        ("Term Loan B, annual amortisation", 0.01, PCT),
        ("Senior notes, turns", 1.0, MULT),
        ("Senior notes, coupon", 0.0725, PCT2),
        ("LTM revenue", 1180.0, MONEY),
        ("Revenue growth", 0.062, PCT),
        ("EBITDA margin expansion, per year", 0.004, PCT),
        ("Capex, % of revenue", 0.038, PCT),
        ("Net working capital, % of revenue", 0.052, PCT),
        ("D&A, % of revenue", 0.031, PCT),
        ("Tax rate", 0.250, PCT),
        ("Exit multiple", 9.5, MULT),
        ("Hold period, years", 5, "0"),
        ("Leverage covenant, max net debt / EBITDA", 6.00, MULT),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value)
        c.number_format = fmt
        c.font = INPUT_FONT

    ws.cell(33, 1, "LTM EBITDA margin").font = BOLD
    ws.cell(33, 2, "=B3/B22").number_format = PCT
    ws.cell(34, 1, "Purchase enterprise value").font = BOLD
    ws.cell(34, 2, "=B3*B4").number_format = MONEY
    ws.cell(35, 1, "Total debt raised").font = BOLD
    ws.cell(35, 2, "=B3*(B14+B17+B20)").number_format = MONEY
    ws.cell(36, 1, "Entry leverage").font = BOLD
    ws.cell(36, 2, "=B35/B3").number_format = MULT
    ws.cell(36, 2).border = TOTAL_BORDER
    return ws


def sources_uses(wb: Workbook):
    ws = wb.create_sheet("SourcesUses")
    widths(ws, a=34, b=14, c=12, d=12)
    title(ws, "Sources and uses of funds")
    header_row(ws, 4, ["Sources", "Amount", "% of total", "x EBITDA"])

    sources = [
        ("Revolver drawn at close", "=0"),
        ("Term Loan A", "=Assumptions!$B$3*Assumptions!$B$14"),
        ("Term Loan B", "=Assumptions!$B$3*Assumptions!$B$17"),
        ("Senior notes", "=Assumptions!$B$3*Assumptions!$B$20"),
        ("Sponsor equity", None),  # the plug, filled in below
    ]
    for i, (label, formula) in enumerate(sources):
        r = 5 + i
        ws.cell(r, 1, label)
        if formula:
            ws.cell(r, 2, formula).number_format = MONEY
        ws.cell(r, 3, f"=IFERROR(B{r}/$B$10,\"\")").number_format = PCT
        ws.cell(r, 4, f"=IFERROR(B{r}/Assumptions!$B$3,\"\")").number_format = MULT
    ws.cell(10, 1, "Total sources").font = BOLD
    ws.cell(10, 2, "=SUM(B5:B9)").number_format = MONEY
    ws.cell(10, 2).font = BOLD
    ws.cell(10, 2).border = TOTAL_BORDER

    header_row(ws, 12, ["Uses", "Amount", "% of total", "x EBITDA"])
    uses = [
        ("Purchase enterprise value", "=Assumptions!$B$34"),
        ("Refinance existing net debt", "=Assumptions!$B$7"),
        ("Transaction fees", "=ROUND(Assumptions!$B$34*Assumptions!$B$5,1)"),
        ("Financing fees", "=ROUND(Assumptions!$B$35*Assumptions!$B$6,1)"),
        ("Cash to balance sheet", "=Assumptions!$B$8"),
    ]
    for i, (label, formula) in enumerate(uses):
        r = 13 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = MONEY
        ws.cell(r, 3, f"=IFERROR(B{r}/$B$18,\"\")").number_format = PCT
        ws.cell(r, 4, f"=IFERROR(B{r}/Assumptions!$B$3,\"\")").number_format = MULT
    ws.cell(18, 1, "Total uses").font = BOLD
    ws.cell(18, 2, "=SUM(B13:B17)").number_format = MONEY
    ws.cell(18, 2).font = BOLD
    ws.cell(18, 2).border = TOTAL_BORDER

    # Sponsor equity is whatever the debt does not fund.
    ws.cell(9, 2, "=B18-SUM(B5:B8)").number_format = MONEY
    ws.cell(9, 2).font = BOLD

    ws.cell(20, 1, "Sources equal uses").font = BOLD
    ws.cell(20, 2, '=IF(ABS(B10-B18)<0.01,"balanced","OUT BY "&TEXT(B10-B18,"#,##0.00"))')
    ws.cell(21, 1, "Sponsor equity, % of capitalisation").font = BOLD
    ws.cell(21, 2, "=B9/B10").number_format = PCT
    ws.cell(22, 1, "Debt / EBITDA at close").font = BOLD
    ws.cell(22, 2, "=SUM(B5:B8)/Assumptions!$B$3").number_format = MULT
    return ws


def model(wb: Workbook):
    """Operating model. Year 0 is LTM; years 1-5 are the hold period.

    The interest line is left empty here and written back by `link_interest`
    once the debt schedule exists — Model and Debt reference each other, and
    resolving that by hand-counting rows in two places is how fixtures acquire
    bugs that then look like engine failures.
    """
    ws = wb.create_sheet("Model")
    widths(ws, a=34)
    for i in range(YEARS + 1):
        ws.column_dimensions[col(FIRST + i)].width = 13
    title(ws, "Operating model")
    ws.freeze_panes = "B5"

    ws.cell(4, 1, "Fiscal year").font = HDR_FONT
    ws.cell(4, 1).fill = HDR_FILL
    for i in range(YEARS + 1):
        c = ws.cell(4, FIRST + i, "LTM" if i == 0 else f'="Year "&{i}')
        c.font = HDR_FONT
        c.fill = HDR_FILL

    g = Grid(ws, FIRST, YEARS + 1, start_row=6)
    prev = lambda i: col(FIRST + i - 1)  # noqa: E731

    g.line("revenue", "Revenue",
           lambda i, c: ("=Assumptions!$B$22" if i == 0
                         else f"=ROUND({prev(i)}{g.R['revenue']}*(1+Assumptions!$B$23),1)"),
           bold=True)
    g.line("margin", "EBITDA margin",
           lambda i, c: ("=Assumptions!$B$33" if i == 0
                         else f"={prev(i)}{g.R['margin']}+Assumptions!$B$24"), PCT)
    g.line("ebitda", "EBITDA",
           lambda i, c: f"=ROUND({c}{g.R['revenue']}*{c}{g.R['margin']},1)", bold=True)
    g.line("da", "Depreciation & amortisation",
           lambda i, c: f"=-ROUND({c}{g.R['revenue']}*Assumptions!$B$27,1)")
    g.line("ebit", "EBIT", lambda i, c: f"={c}{g.R['ebitda']}+{c}{g.R['da']}", bold=True)
    # Interest comes back from the debt schedule — see link_interest.
    g.line("interest", "Interest expense")
    g.line("ebt", "Profit before tax",
           lambda i, c: f"={c}{g.R['ebit']}+{c}{g.R['interest']}")
    g.line("tax", "Taxes",
           lambda i, c: f"=-MAX(0,ROUND({c}{g.R['ebt']}*Assumptions!$B$28,1))")
    g.line("ni", "Net income", lambda i, c: f"={c}{g.R['ebt']}+{c}{g.R['tax']}", bold=True)
    g.gap()
    g.line("addback", "Add back D&A", lambda i, c: f"=-{c}{g.R['da']}")
    g.line("capex", "Capital expenditure",
           lambda i, c: f"=-ROUND({c}{g.R['revenue']}*Assumptions!$B$25,1)")
    g.line("nwc", "NWC balance",
           lambda i, c: f"=ROUND({c}{g.R['revenue']}*Assumptions!$B$26,1)")
    g.line("dnwc", "(Increase) in NWC",
           lambda i, c: ("=0" if i == 0
                         else f"=-({c}{g.R['nwc']}-{prev(i)}{g.R['nwc']})"))
    g.line("fcf", "Free cash flow before debt service",
           lambda i, c: (f"={c}{g.R['ni']}+{c}{g.R['addback']}"
                         f"+{c}{g.R['capex']}+{c}{g.R['dnwc']}"), bold=True)
    g.render()
    for i in range(YEARS + 1):
        ws.cell(g.R["fcf"], FIRST + i).border = TOTAL_BORDER

    r = g.r + 1
    ws.cell(r, 1, "Cumulative FCF over hold").font = BOLD
    ws.cell(r, 2, f"=SUM(C{g.R['fcf']}:{col(FIRST + YEARS)}{g.R['fcf']})").number_format = MONEY
    ws.cell(r + 1, 1, "EBITDA CAGR").font = BOLD
    ws.cell(r + 1, 2,
            f"=({col(FIRST + YEARS)}{g.R['ebitda']}/B{g.R['ebitda']})^(1/{YEARS})-1"
            ).number_format = PCT
    return ws, g


def debt(wb: Workbook, model_rows: dict):
    ws = wb.create_sheet("Debt")
    widths(ws, a=36)
    for i in range(YEARS):
        ws.column_dimensions[col(FIRST + i)].width = 13
    title(ws, "Debt schedule and cash sweep")
    ws.freeze_panes = "B5"

    ws.cell(4, 1, "Fiscal year").font = HDR_FONT
    ws.cell(4, 1).fill = HDR_FILL
    for i in range(YEARS):
        c = ws.cell(4, FIRST + i, f'="Year "&{i + 1}')
        c.font = HDR_FONT
        c.fill = HDR_FILL

    g = Grid(ws, FIRST, YEARS, start_row=6)
    prev = lambda i: col(FIRST + i - 1)  # noqa: E731
    # Model column for debt year i (0-based) — Model column B is year 0.
    mcol = lambda i: col(FIRST + i + 1)  # noqa: E731

    g.line("cash_begin", "Cash, beginning",
           lambda i, c: ("=SourcesUses!$B$17" if i == 0
                         else f"={prev(i)}{g.R['cash_end']}"))
    g.line("rev_begin", "Revolver, beginning",
           lambda i, c: ("=SourcesUses!$B$5" if i == 0 else f"={prev(i)}{g.R['rev_end']}"))
    g.line("tla_begin", "Term Loan A, beginning",
           lambda i, c: ("=SourcesUses!$B$6" if i == 0 else f"={prev(i)}{g.R['tla_end']}"))
    g.line("tlb_begin", "Term Loan B, beginning",
           lambda i, c: ("=SourcesUses!$B$7" if i == 0 else f"={prev(i)}{g.R['tlb_end']}"))
    g.line("notes_begin", "Senior notes, beginning",
           lambda i, c: ("=SourcesUses!$B$8" if i == 0 else f"={prev(i)}{g.R['notes_end']}"))
    g.gap()

    ws.cell(g.r, 1, "Interest and fees").font = BOLD
    ws.cell(g.r, 1).fill = SUB_FILL
    g.gap()
    g.line("int_rev", "Revolver interest",
           lambda i, c: f"=ROUND({c}{g.R['rev_begin']}*(Assumptions!$B$11+Assumptions!$B$12),1)",
           indent=1)
    g.line("fee_rev", "Revolver commitment fee",
           lambda i, c: (f"=ROUND((Assumptions!$B$10-{c}{g.R['rev_begin']})"
                         f"*Assumptions!$B$13,1)"), indent=1)
    g.line("int_tla", "Term Loan A interest",
           lambda i, c: f"=ROUND({c}{g.R['tla_begin']}*(Assumptions!$B$11+Assumptions!$B$15),1)",
           indent=1)
    g.line("int_tlb", "Term Loan B interest",
           lambda i, c: f"=ROUND({c}{g.R['tlb_begin']}*(Assumptions!$B$11+Assumptions!$B$18),1)",
           indent=1)
    g.line("int_notes", "Senior notes interest",
           lambda i, c: f"=ROUND({c}{g.R['notes_begin']}*Assumptions!$B$21,1)", indent=1)
    g.line("interest_total", "Total interest and fees",
           lambda i, c: f"=SUM({c}{g.R['int_rev']}:{c}{g.R['int_notes']})", bold=True)
    g.gap()

    ws.cell(g.r, 1, "Debt service").font = BOLD
    ws.cell(g.r, 1).fill = SUB_FILL
    g.gap()
    g.line("mand_tla", "Mandatory amortisation — TLA",
           lambda i, c: (f"=MIN({c}{g.R['tla_begin']},"
                         f"ROUND(SourcesUses!$B$6*Assumptions!$B$16,1))"), indent=1)
    g.line("mand_tlb", "Mandatory amortisation — TLB",
           lambda i, c: (f"=MIN({c}{g.R['tlb_begin']},"
                         f"ROUND(SourcesUses!$B$7*Assumptions!$B$19,1))"), indent=1)
    g.line("mand_total", "Total mandatory amortisation",
           lambda i, c: f"={c}{g.R['mand_tla']}+{c}{g.R['mand_tlb']}")
    g.line("fcf", "Free cash flow after interest and tax",
           lambda i, c: f"=Model!{mcol(i)}{model_rows['fcf']}")
    g.line("cash_pre", "Cash before sweep",
           lambda i, c: (f"={c}{g.R['cash_begin']}+{c}{g.R['fcf']}-{c}{g.R['mand_total']}"))
    g.line("sweep", "Cash available to sweep",
           lambda i, c: (f"=MAX(0,ROUND(Assumptions!$B$9*"
                         f"({c}{g.R['cash_pre']}-Assumptions!$B$8),1))"), bold=True)
    g.gap()

    # The waterfall: revolver first, then TLA, then TLB. Each tranche only sees
    # what the ones above it left behind.
    g.line("pre_rev", "Optional prepayment — revolver",
           lambda i, c: f"=MIN({c}{g.R['rev_begin']},{c}{g.R['sweep']})", indent=1)
    g.line("pre_tla", "Optional prepayment — TLA",
           lambda i, c: (f"=MIN({c}{g.R['tla_begin']}-{c}{g.R['mand_tla']},"
                         f"MAX(0,{c}{g.R['sweep']}-{c}{g.R['pre_rev']}))"), indent=1)
    g.line("pre_tlb", "Optional prepayment — TLB",
           lambda i, c: (f"=MIN({c}{g.R['tlb_begin']}-{c}{g.R['mand_tlb']},"
                         f"MAX(0,{c}{g.R['sweep']}-{c}{g.R['pre_rev']}-{c}{g.R['pre_tla']}))"),
           indent=1)
    g.line("pre_total", "Total optional prepayment",
           lambda i, c: (f"=SUM({c}{g.R['pre_rev']}:{c}{g.R['pre_tlb']})"), bold=True)
    g.gap()

    g.line("rev_end", "Revolver, ending",
           lambda i, c: f"={c}{g.R['rev_begin']}-{c}{g.R['pre_rev']}")
    g.line("tla_end", "Term Loan A, ending",
           lambda i, c: f"={c}{g.R['tla_begin']}-{c}{g.R['mand_tla']}-{c}{g.R['pre_tla']}")
    g.line("tlb_end", "Term Loan B, ending",
           lambda i, c: f"={c}{g.R['tlb_begin']}-{c}{g.R['mand_tlb']}-{c}{g.R['pre_tlb']}")
    g.line("notes_end", "Senior notes, ending",
           lambda i, c: f"={c}{g.R['notes_begin']}")
    g.line("cash_end", "Cash, ending",
           lambda i, c: f"={c}{g.R['cash_pre']}-{c}{g.R['pre_total']}", bold=True)
    g.line("total_debt", "Total debt, ending",
           lambda i, c: f"=SUM({c}{g.R['rev_end']}:{c}{g.R['notes_end']})", bold=True)
    g.line("net_debt", "Net debt, ending",
           lambda i, c: f"={c}{g.R['total_debt']}-{c}{g.R['cash_end']}", bold=True)
    g.gap()

    g.line("leverage", "Net debt / EBITDA",
           lambda i, c: f"={c}{g.R['net_debt']}/Model!{mcol(i)}{model_rows['ebitda']}", MULT)
    g.line("coverage", "EBITDA / interest",
           lambda i, c: (f"=IFERROR(Model!{mcol(i)}{model_rows['ebitda']}"
                         f"/{c}{g.R['interest_total']},\"n/m\")"), MULT)
    g.line("covenant", "Covenant test",
           lambda i, c: (f'=IF({c}{g.R["leverage"]}>Assumptions!$B$31,'
                         f'"BREACH — "&TEXT({c}{g.R["leverage"]},"0.00")&"x","ok")'), "General")
    g.line("paydown", "Cumulative debt paydown",
           lambda i, c: (f"=SourcesUses!$B$10-SourcesUses!$B$9-{c}{g.R['total_debt']}"))

    g.render()

    r = g.r + 1
    ws.cell(r, 1, "Years in covenant breach").font = BOLD
    ws.cell(r, 2, f'=COUNTIF({g.span("covenant")},"BREACH*")').number_format = "0"
    ws.cell(r + 1, 1, "Revolver ever drawn").font = BOLD
    ws.cell(r + 1, 2, f'=IF(MAX({g.span("rev_end")})>0,"yes","no")')
    ws.cell(r + 2, 1, "Total interest paid over hold").font = BOLD
    ws.cell(r + 2, 2, f"=SUM({g.span('interest_total')})").number_format = MONEY
    return ws, g


def returns(wb: Workbook, mg: Grid, dg: Grid):
    ws = wb.create_sheet("Returns")
    widths(ws, a=36, b=14, c=13, d=13, e=13, f=13, g=13)
    title(ws, "Sponsor returns")

    exit_col = col(FIRST + YEARS)      # Model year 5
    debt_exit = col(FIRST + YEARS - 1)  # Debt year 5

    rows = [
        ("exit_ebitda", "Exit year EBITDA", f"=Model!{exit_col}{mg.R['ebitda']}", MONEY),
        ("exit_ev", "Exit enterprise value", "=B5*Assumptions!$B$29", MONEY),
        ("exit_nd", "Less: net debt at exit", f"=-Debt!{debt_exit}{dg.R['net_debt']}", MONEY),
        ("proceeds", "Equity proceeds", "=B6+B7", MONEY),
        ("equity_in", "Sponsor equity invested", "=SourcesUses!$B$9", MONEY),
        ("moic", "Multiple of invested capital", "=B8/B9", MULT),
        ("irr_alg", "IRR (algebraic)", "=B10^(1/Assumptions!$B$30)-1", PCT2),
    ]
    for i, (_key, label, formula, fmt) in enumerate(rows):
        r = 5 + i
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 2, formula).number_format = fmt
    ws.cell(11, 2).border = TOTAL_BORDER

    ws.cell(13, 1, "Sponsor cash flows").font = BOLD
    ws.cell(13, 1).fill = SUB_FILL
    for i in range(YEARS + 1):
        ws.cell(13, FIRST + i, f'="Year "&{i}').font = HDR_FONT
        ws.cell(13, FIRST + i).fill = HDR_FILL
    for i in range(YEARS + 1):
        if i == 0:
            f = "=-B9"
        elif i == YEARS:
            f = "=B8"
        else:
            f = "=0"
        ws.cell(14, FIRST + i, f).number_format = MONEY
    ws.cell(14, 1, "Equity cash flow")

    flows = f"B14:{col(FIRST + YEARS)}14"
    ws.cell(16, 1, "IRR (function)").font = BOLD
    ws.cell(16, 2, f"=IRR({flows})").number_format = PCT2
    ws.cell(17, 1, "IRR methods agree").font = BOLD
    ws.cell(17, 2, '=IF(ABS(B16-B11)<0.0001,"ties","OUT BY "&TEXT(B16-B11,"0.0000%"))')
    ws.cell(18, 1, "NPV of equity flows at 20%").font = BOLD
    ws.cell(18, 2, f"=NPV(0.2,C14:{col(FIRST + YEARS)}14)+B14").number_format = MONEY

    ws.cell(20, 1, "Value creation attribution").font = BOLD
    ws.cell(20, 1).fill = SUB_FILL
    attribution = [
        ("EBITDA growth", "=(B5-Assumptions!$B$3)*Assumptions!$B$4"),
        ("Multiple expansion", "=(Assumptions!$B$29-Assumptions!$B$4)*B5"),
        ("Debt paydown and cash build",
         f"=(SUM(SourcesUses!$B$5:$B$8)-Assumptions!$B$8)-Debt!{debt_exit}{dg.R['net_debt']}"),
    ]
    for i, (label, formula) in enumerate(attribution):
        r = 21 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = MONEY
        ws.cell(r, 3, "=IFERROR(B%d/$B$24,\"\")" % r).number_format = PCT
    ws.cell(24, 1, "Total value created").font = BOLD
    ws.cell(24, 2, "=SUM(B21:B23)").number_format = MONEY
    ws.cell(25, 1, "Reconciles to equity gain").font = BOLD
    ws.cell(25, 2, '=IF(ABS(B24-(B8-B9))<1,"ties","OUT BY "&TEXT(B24-(B8-B9),"#,##0.0"))')
    return ws


def matrix(wb: Workbook, mg: Grid, dg: Grid):
    ws = wb.create_sheet("ReturnsMatrix")
    widths(ws, a=16, b=13, c=13, d=13)
    title(ws, "IRR sensitivity — exit multiple vs hold period")
    note(ws, 2, "Reads EBITDA and net debt out of the schedules by year index.")

    top = 5
    mults = 7
    holds = [3, 4, 5]
    ws.cell(top, 1, "Multiple \\ yr").font = HDR_FONT
    ws.cell(top, 1).fill = HDR_FILL
    for j, h in enumerate(holds):
        c = ws.cell(top, 2 + j, h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.number_format = "0"
    for i in range(mults):
        c = ws.cell(top + 1 + i, 1, f"=Assumptions!$B$29+{(i - (mults - 1) // 2) * 0.5}")
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.number_format = MULT

    ebitda_band = (f"Model!$B${mg.R['ebitda']}:"
                   f"${col(FIRST + YEARS)}${mg.R['ebitda']}")
    nd_band = (f"Debt!$B${dg.R['net_debt']}:"
               f"${col(FIRST + YEARS - 1)}${dg.R['net_debt']}")
    for i in range(mults):
        r = top + 1 + i
        for j in range(len(holds)):
            hc = f"{col(2 + j)}${top}"
            cell = ws.cell(
                r, 2 + j,
                f"=IFERROR(((INDEX({ebitda_band},1,{hc}+1)*$A{r}"
                f"-INDEX({nd_band},1,{hc}))/SourcesUses!$B$9)^(1/{hc})-1,\"n/m\")",
            )
            cell.number_format = PCT2

    base = top + mults + 2
    grid = f"B{top + 1}:{col(1 + len(holds))}{top + mults}"
    ws.cell(base, 1, "Base case cell").font = BOLD
    ws.cell(base, 2, f"={col(2 + len(holds) - 1)}{top + 1 + (mults - 1) // 2}").number_format = PCT2
    ws.cell(base + 1, 1, "Returns sheet IRR").font = BOLD
    ws.cell(base + 1, 2, "=Returns!B11").number_format = PCT2
    ws.cell(base + 2, 1, "Matrix ties to model").font = BOLD
    ws.cell(base + 2, 2,
            f'=IF(ABS(B{base}-B{base + 1})<0.0005,"ties",'
            f'"OUT BY "&TEXT(B{base}-B{base + 1},"0.00%"))')
    ws.cell(base + 4, 1, "Outcomes above 25% IRR").font = BOLD
    ws.cell(base + 4, 2, f'=COUNTIF({grid},">0.25")').number_format = "0"
    ws.cell(base + 5, 1, "Best case").font = BOLD
    ws.cell(base + 5, 2, f"=MAX({grid})").number_format = PCT2
    ws.cell(base + 6, 1, "Worst case").font = BOLD
    ws.cell(base + 6, 2, f"=MIN({grid})").number_format = PCT2
    return ws


def link_interest(ws, mg: Grid, dg: Grid) -> None:
    """Close the Model → Debt → Model loop, now that both row maps are known."""
    for i in range(YEARS + 1):
        cell = ws.cell(mg.R["interest"], FIRST + i)
        cell.value = "=0" if i == 0 else f"=-Debt!{col(FIRST + i - 1)}{dg.R['interest_total']}"
        cell.number_format = MONEY


def build() -> Workbook:
    wb = Workbook()
    assumptions(wb)
    sources_uses(wb)
    ws_model, mg = model(wb)
    _, dg = debt(wb, mg.R)
    link_interest(ws_model, mg, dg)
    returns(wb, mg, dg)
    matrix(wb, mg, dg)
    return wb
