"""M10 — The adversarial sheet: semantics, errors and things that should refuse.

Every other model in this corpus is a model. This one is a test, written the way
a careful reviewer probes a new engine: one labelled case per row, each isolating
a single piece of Excel behaviour that is easy to get plausibly wrong.

Three groups, and they are judged differently:

  semantics    `-2^2` is 4, `"5"+1` is 6 but `"5"=5` is FALSE, `MOD(-3,2)` is 1,
               `ROUND(2.675,2)` is 2.68. These must match the oracle exactly.

  errors       `1/0` is #DIV/0! and a missing VLOOKUP is #N/A. These are
               *computed values*, not failures — an engine that renders ⚠ here
               is as wrong as one that renders a number.

  refusals     `INDIRECT`, `OFFSET`, a circular reference, an approximate lookup
               over unsorted data. Here ⚠ is the correct answer and a number is
               the failure. The point of the sheet is that these three groups
               must not leak into one another.

RAND and RANDBETWEEN are deliberately absent: they cannot agree with an oracle
by construction, so including them would only add noise. TODAY and NOW are
present, because a preview has to do *something* defensible with them and that
something is worth seeing in the report.
"""

from __future__ import annotations

from openpyxl import Workbook

from common import (
    BOLD, DATE, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY, MONEY2, NUM,
    SUB_FILL, TOP_BORDER, col, header_row, note, title, widths,
)

# (group, label, formula, what Excel does)
CASES = [
    # ---- operator precedence and associativity ----
    ("operators", "Unary minus binds tighter than power", "=-2^2",
     "4 — Excel is the odd one out here; most languages give -4"),
    ("operators", "Power is left-associative", "=2^3^2", "64, not 512"),
    ("operators", "Percent is a postfix operator", "=50%*200", "100"),
    ("operators", "Percent binds tighter than power", "=2^2%", "1.0139… (2^0.02)"),
    ("operators", "Unary minus chains", "=--5", "5"),
    ("operators", "Comparison after concatenation", '="a"&"b"="ab"', "TRUE"),
    ("operators", "Multiplication before addition", "=2+3*4", "14"),
    ("operators", "Parenthesised negation", "=(-2)^2", "4"),
    ("operators", "Range then intersection", "=SUM(H4:H6 H5:H8)",
     "intersection of two ranges — the space operator"),

    # ---- coercion ----
    ("coercion", "Text that looks numeric, in arithmetic", '="5"+1', "6"),
    ("coercion", "Text that looks numeric, in comparison", '="5"=5',
     "FALSE — arithmetic coerces, comparison does not"),
    ("coercion", "Boolean in arithmetic", "=TRUE+1", "2"),
    ("coercion", "Boolean compared to number", "=TRUE=1", "FALSE"),
    ("coercion", "Text ranks above number", '=("a">1)', "TRUE"),
    ("coercion", "Boolean ranks above text", '=(TRUE>"z")', "TRUE"),
    ("coercion", "Case-insensitive text comparison", '="a"="A"', "TRUE"),
    ("coercion", "EXACT is case-sensitive", '=EXACT("a","A")', "FALSE"),
    ("coercion", "Numeric text in a range is ignored by SUM", "=SUM(H4:H15)",
     "text and booleans in a range do not contribute"),
    ("coercion", "Numeric text as a direct argument counts", '=SUM("3",1)', "4"),
    ("coercion", "Concatenation stringifies numbers", '=1/3&""',
     "0.333333333333333 — 15 significant digits"),
    ("coercion", "Large number stringified", "=1234567890123456789&\"\"",
     "1.23456789012346E+18"),

    # ---- blanks ----
    ("blanks", "Reference to an empty cell", "=Z1", "0"),
    ("blanks", "Empty cell equals zero", "=Z1=0", "TRUE"),
    ("blanks", "Empty cell equals empty string", '=Z1=""',
     "TRUE — a blank satisfies both, which no other value does"),
    ("blanks", "ISBLANK on an empty cell", "=ISBLANK(Z1)", "TRUE"),
    ("blanks", "ISBLANK on a formula returning empty text", '=ISBLANK(D2)',
     "FALSE — a formula result is never blank"),
    ("blanks", "Empty string is not zero", '=""=0', "FALSE"),
    ("blanks", "COUNT over a mixed range", "=COUNT(H4:H15)", "numbers only"),
    ("blanks", "COUNTA over a mixed range", "=COUNTA(H4:H15)", "anything non-empty"),
    ("blanks", "COUNTBLANK over a mixed range", "=COUNTBLANK(H4:H15)", "truly empty only"),
    ("blanks", "AVERAGE skips text and blanks", "=AVERAGE(H4:H15)",
     "denominator is COUNT, not COUNTA"),

    # ---- rounding and numerics ----
    ("numerics", "Round half away from zero", "=ROUND(2.5,0)", "3"),
    ("numerics", "Round half away from zero, negative", "=ROUND(-2.5,0)", "-3"),
    ("numerics", "Binary representation does not decide it", "=ROUND(2.675,2)",
     "2.68 — the decimal value rounds up even though the double is below it"),
    ("numerics", "Round to tens", "=ROUND(1234.5678,-2)", "1200"),
    ("numerics", "ROUNDDOWN truncates toward zero", "=ROUNDDOWN(-2.9,0)", "-2"),
    ("numerics", "INT floors toward negative infinity", "=INT(-2.5)", "-3"),
    ("numerics", "TRUNC truncates toward zero", "=TRUNC(-2.5)", "-2"),
    ("numerics", "MOD takes the sign of the divisor", "=MOD(-3,2)", "1"),
    ("numerics", "MOD with a negative divisor", "=MOD(3,-2)", "-1"),
    ("numerics", "Floating point addition", "=0.1+0.2=0.3",
     "FALSE in IEEE 754 — Excel does not hide this in a comparison"),
    ("numerics", "But the displayed sum is clean", "=0.1+0.2", "0.3"),
    ("numerics", "Overflow", "=1E+308*10", "#NUM!"),
    ("numerics", "Underflow to zero", "=1E-300/1E+30", "0"),
    ("numerics", "Fifteen significant digits", "=1234567890123456+1",
     "Excel keeps 15 digits, so this does not change"),

    # ---- dates ----
    ("dates", "Serial 1 is 1 January 1900", '=TEXT(1,"yyyy-mm-dd")', "1900-01-01"),
    ("dates", "Serial 60 is the phantom leap day", '=TEXT(60,"yyyy-mm-dd")',
     "1900-02-29 — a date that never existed, kept for Lotus compatibility"),
    ("dates", "Serial 61 is 1 March 1900", '=TEXT(61,"yyyy-mm-dd")', "1900-03-01"),
    ("dates", "DATE rolls over month ends", '=TEXT(DATE(2026,13,1),"yyyy-mm-dd")',
     "2027-01-01"),
    ("dates", "DATE rolls over day zero", '=TEXT(DATE(2026,3,0),"yyyy-mm-dd")',
     "2026-02-28"),
    ("dates", "EDATE clamps to the shorter month",
     '=TEXT(EDATE(DATE(2026,1,31),1),"yyyy-mm-dd")', "2026-02-28"),
    ("dates", "EOMONTH of a leap February",
     '=TEXT(EOMONTH(DATE(2024,2,1),0),"yyyy-mm-dd")', "2024-02-29"),
    ("dates", "Date arithmetic is plain subtraction",
     "=DATE(2026,12,25)-DATE(2026,1,1)", "358"),
    ("dates", "Time is the fractional part", "=TIME(18,30,0)*24", "18.5"),
    ("dates", "YEARFRAC on the actual/actual basis",
     "=YEARFRAC(DATE(2026,1,1),DATE(2026,7,1),1)", "about 0.4959"),
    ("dates", "WEEKDAY with the default basis", "=WEEKDAY(DATE(2026,7,25))",
     "7 — Saturday, Sunday=1"),
    ("dates", "NETWORKDAYS excludes weekends",
     "=NETWORKDAYS(DATE(2026,7,1),DATE(2026,7,31))", "23"),

    # ---- errors, which are values ----
    ("errors", "Division by zero", "=1/0", "#DIV/0!"),
    ("errors", "Caught division by zero", '=IFERROR(1/0,"caught")', '"caught"'),
    ("errors", "Errors propagate through arithmetic", "=1/0+1", "#DIV/0!"),
    ("errors", "Errors propagate through SUM", "=SUM(H4:H15)+H12",
     "#DIV/0! — H12 is an error"),
    ("errors", "Text where a number is required", '=1+"apple"', "#VALUE!"),
    ("errors", "Lookup miss", '=VLOOKUP("nothing",K4:L8,2,FALSE)', "#N/A"),
    ("errors", "IFNA catches only #N/A",
     '=IFNA(VLOOKUP("nothing",K4:L8,2,FALSE),"missing")', '"missing"'),
    ("errors", "IFNA does not catch #DIV/0!", '=IFNA(1/0,"missing")', "#DIV/0!"),
    ("errors", "Square root of a negative", "=SQRT(-1)", "#NUM!"),
    ("errors", "Log of zero", "=LOG(0)", "#NUM!"),
    ("errors", "INDEX past the end of the range", "=INDEX(H4:H6,9)", "#REF!"),
    ("errors", "CHOOSE with an out-of-range index", '=CHOOSE(9,"a","b")', "#VALUE!"),
    ("errors", "Unknown name", "=NOTAFUNCTION(1)", "#NAME?"),
    ("errors", "Error inside a false branch is never evaluated",
     '=IF(TRUE,"safe",1/0)', '"safe" — IF is lazy'),
    ("errors", "Error inside a true branch is evaluated", '=IF(FALSE,"safe",1/0)',
     "#DIV/0!"),
    ("errors", "ISERROR sees it", "=ISERROR(1/0)", "TRUE"),
    ("errors", "ERROR.TYPE numbers it", "=ERROR.TYPE(1/0)", "2"),
    ("errors", "N/A propagates through AVERAGE", "=AVERAGE(H4:H6,NA())", "#N/A"),
    ("errors", "AGGREGATE could skip it, but is not implemented everywhere",
     "=SUM(H4:H6)", "control case — plain SUM of clean numbers"),

    # ---- lookups ----
    ("lookups", "Exact match finds it", '=VLOOKUP(30,K4:L8,2,FALSE)', '"thirty"'),
    ("lookups", "Approximate match over a SORTED table",
     "=VLOOKUP(25,M4:N8,2,TRUE)", "the 20 row — sorted, so well defined"),
    ("lookups", "Approximate match over an UNSORTED table",
     "=VLOOKUP(25,K4:L8,2,TRUE)",
     "Excel answers from its binary search's probe order — arbitrary, and a "
     "defensible thing to refuse"),
    ("lookups", "MATCH exact", "=MATCH(30,K4:K8,0)", "4"),
    ("lookups", "MATCH descending over unsorted data", "=MATCH(25,K4:K8,-1)",
     "same arbitrariness as above"),
    ("lookups", "HLOOKUP across a row", "=HLOOKUP(20,H4:J4,1,FALSE)", "#N/A"),
    ("lookups", "INDEX and MATCH together",
     "=INDEX(L4:L8,MATCH(70,K4:K8,0))", '"seventy"'),
    ("lookups", "XLOOKUP with a not-found fallback",
     '=XLOOKUP(99,K4:K8,L4:L8,"none")', '"none"'),
    ("lookups", "Whole-column reference", "=SUM(H:H)", "sums the data block"),
    ("lookups", "Whole-row reference", "=COUNT(4:4)", "counts numbers in row 4"),

    # ---- text ----
    ("text", "LEN counts characters", '=LEN("naïve café")', "10"),
    ("text", "Non-ASCII survives round trip", '=UPPER("straße")', "STRASSE or STRAßE"),
    ("text", "MID past the end returns empty", '=MID("abc",9,2)', '""'),
    ("text", "LEFT with a zero count", '=LEFT("abc",0)', '""'),
    ("text", "SUBSTITUTE replaces the nth occurrence",
     '=SUBSTITUTE("a-b-c","-","+",2)', '"a-b+c"'),
    ("text", "TEXT with a thousands format", '=TEXT(1234.5,"#,##0.00")', '"1,234.50"'),
    ("text", "TEXT of a negative in a bracket format",
     '=TEXT(-1234.5,"#,##0;(#,##0)")', '"(1,235)"'),
    ("text", "TEXTJOIN skips empty when asked",
     '=TEXTJOIN("|",TRUE,"a","","b")', '"a|b"'),
    ("text", "TEXTJOIN keeps empty when told",
     '=TEXTJOIN("|",FALSE,"a","","b")', '"a||b"'),
    ("text", "VALUE parses a formatted number", '=VALUE("1,234.5")', "1234.5"),
    ("text", "TRIM collapses internal runs", '=TRIM("  a   b  ")', '"a b"'),
    ("text", "REPT builds a long string", '=LEN(REPT("ab",200))', "400"),
    ("text", "CONCAT of many arguments",
     '=LEN("x"&"y"&"z"&REPT("-",250))', "253"),
    ("text", "CHAR and CODE round trip", "=CODE(CHAR(65))", "65"),
    ("text", "Text compared to a number is greater", '=IF("10">9,"text wins","number wins")',
     '"text wins" — type rank beats numeric value'),

    # ---- things that should refuse ----
    ("refusals", "INDIRECT builds a reference at run time", '=INDIRECT("H4")',
     "10 in Excel; a static dependency graph cannot express it"),
    ("refusals", "OFFSET moves a reference at run time", "=OFFSET(H4,1,0)",
     "20 in Excel; same problem"),
    ("refusals", "Volatile TODAY", "=TODAY()", "changes daily by design"),
    ("refusals", "Volatile NOW truncated", "=INT(NOW())", "same as TODAY"),
    ("refusals", "CELL asks about the workbook", '=CELL("address",H4)', "$H$4"),
    ("refusals", "INFO asks about the environment", '=INFO("numfile")',
     "environment-dependent"),

    # ---- array-shaped ----
    ("arrays", "Array constant in SUM", "=SUM({1,2,3})", "6"),
    ("arrays", "Two-dimensional array constant", "=SUM({1,2;3,4})", "10"),
    ("arrays", "SUMPRODUCT over two ranges",
     "=SUMPRODUCT(H4:H6,H13:H15)", "element-wise then summed"),
    ("arrays", "SUMPRODUCT with a condition",
     '=SUMPRODUCT((I4:I8="beta")*H4:H8)', "20"),
    ("arrays", "SUMPRODUCT with mismatched shapes",
     "=SUMPRODUCT(H4:H6,H4:H8)", "#VALUE!"),
    ("arrays", "Nested function depth",
     "=IF(1,IF(1,IF(1,IF(1,IF(1,IF(1,IF(1,IF(1,IF(1,IF(1,"
     '"ten deep",0),0),0),0),0),0),0),0),0),0)', '"ten deep"'),
]


def traps(wb: Workbook):
    ws = wb.active
    ws.title = "Traps"
    widths(ws, a=13, b=44, c=48, d=16, h=10, i=12, j=10, k=10, l=12, m=10, n=12)
    title(ws, "Semantics, errors and refusals")
    note(ws, 2, "Column D is live. Column C says what Excel does.")
    # D2 holds a formula that returns an empty string — referenced by the
    # ISBLANK case, which must distinguish it from a genuinely empty cell.
    ws.cell(2, 4, '=IF(TRUE,"","x")')

    # Scratch data the probes read. Deliberately mixed: numbers, text, a
    # boolean, a numeric-looking string, a hole, and an error.
    ws.cell(3, 8, "Data").font = BOLD
    ws.cell(3, 8).fill = SUB_FILL
    data = [10, 20, 30, None, "text", True, "5", 0, "=1/0", 40, -15, 2.5]
    for i, v in enumerate(data):
        if v is None:
            continue  # a real hole, not an empty string
        c = ws.cell(4 + i, 8, v)
        c.font = INPUT_FONT
    keys = ["alpha", "beta", "gamma", "", "delta"]
    for i, k in enumerate(keys):
        if k:
            ws.cell(4 + i, 9, k)

    ws.cell(3, 11, "Unsorted lookup table").font = BOLD
    for i, (n, label) in enumerate([(50, "fifty"), (10, "ten"), (90, "ninety"),
                                    (30, "thirty"), (70, "seventy")]):
        ws.cell(4 + i, 11, n).number_format = INT
        ws.cell(4 + i, 12, label)

    ws.cell(3, 13, "Sorted lookup table").font = BOLD
    for i, (n, label) in enumerate([(0, "zero"), (10, "ten"), (20, "twenty"),
                                    (30, "thirty"), (40, "forty")]):
        ws.cell(4 + i, 13, n).number_format = INT
        ws.cell(4 + i, 14, label)

    header_row(ws, 4, ["Group", "Case", "Excel's answer", "Live"], start_col=1)
    for i, (group, label, formula, expected) in enumerate(CASES):
        r = 5 + i
        ws.cell(r, 1, group)
        ws.cell(r, 2, label)
        ws.cell(r, 3, expected)
        ws.cell(r, 4, formula)
    return ws


def circular(wb: Workbook):
    """Genuine cycles. The only correct render here is a refusal."""
    ws = wb.create_sheet("Circular")
    widths(ws, a=34, b=16, c=48)
    title(ws, "Circular references")
    note(ws, 2, "Excel without iterative calculation shows a warning and zeroes. "
                "Anything that renders a confident number here is wrong.")

    header_row(ws, 4, ["Case", "Live", "Note"])
    ws.cell(5, 1, "Two-cell cycle, first leg")
    ws.cell(5, 2, "=B6+1")
    ws.cell(5, 3, "B5 needs B6, B6 needs B5")
    ws.cell(6, 1, "Two-cell cycle, second leg")
    ws.cell(6, 2, "=B5+1")

    ws.cell(8, 1, "Self reference")
    ws.cell(8, 2, "=B8+1")
    ws.cell(8, 3, "the shortest possible cycle")

    ws.cell(10, 1, "Three-cell cycle")
    ws.cell(10, 2, "=B11*2")
    ws.cell(11, 2, "=B12+3")
    ws.cell(12, 2, "=B10/2")
    ws.cell(10, 3, "B10 → B11 → B12 → B10")

    ws.cell(14, 1, "Interest on the average balance")
    ws.cell(14, 3, "the classic accidental circularity in a debt schedule")
    ws.cell(15, 1, "Opening balance")
    ws.cell(15, 2, 1000.0).number_format = MONEY
    ws.cell(16, 1, "Interest")
    ws.cell(16, 2, "=ROUND(AVERAGE(B15,B18)*0.08,2)").number_format = MONEY
    ws.cell(17, 1, "Repayment")
    ws.cell(17, 2, 200.0).number_format = MONEY
    ws.cell(18, 1, "Closing balance")
    ws.cell(18, 2, "=B15+B16-B17").number_format = MONEY

    ws.cell(20, 1, "Clean cell downstream of the cycle").font = BOLD
    ws.cell(20, 2, "=B18*2")
    ws.cell(20, 3, "must not be rendered as a number either — "
                   "an uncomputable input poisons its dependents")
    ws.cell(22, 1, "Clean cell with no cycle in its history").font = BOLD
    ws.cell(22, 2, "=B15*2")
    ws.cell(22, 3, "must still compute — poisoning is not contagion by proximity")
    return ws


def hardcoded(wb: Workbook):
    """A total that was typed rather than summed. The audit case."""
    ws = wb.create_sheet("Hardcoded")
    widths(ws, a=28, b=14, c=14, d=14, e=14, f=14)
    title(ws, "A model with a typed-in total")
    note(ws, 2, "Nothing on the rendered sheet distinguishes D9 from its neighbours.")

    header_row(ws, 4, ["Region", "Q1", "Q2", "Q3", "Q4", "FY"])
    rows = [
        ("Americas", 1240.0, 1318.0, 1402.0, 1495.0),
        ("EMEA", 880.0, 902.0, 948.0, 1011.0),
        ("APAC", 512.0, 561.0, 604.0, 668.0),
        ("Other", 96.0, 101.0, 108.0, 119.0),
    ]
    for i, (label, *vals) in enumerate(rows):
        r = 5 + i
        ws.cell(r, 1, label)
        for j, v in enumerate(vals):
            ws.cell(r, 2 + j, v).number_format = MONEY
        ws.cell(r, 6, f"=SUM(B{r}:E{r})").number_format = MONEY

    ws.cell(9, 1, "Total").font = BOLD
    for j in range(4):
        c = col(2 + j)
        cell = ws.cell(9, 2 + j)
        # Q3 is typed in and 60 too high: 1402+948+604+108 is 3062, not 3122.
        # It is in the right place, formatted like its neighbours, and plausible.
        cell.value = 3122.0 if j == 2 else f"=SUM({c}5:{c}8)"
        cell.number_format = MONEY
        cell.font = BOLD
        cell.border = TOP_BORDER
    ws.cell(9, 6, "=SUM(B9:E9)").number_format = MONEY
    ws.cell(9, 6).font = BOLD
    ws.cell(9, 6).border = TOP_BORDER

    ws.cell(11, 1, "Cross-foot check").font = BOLD
    ws.cell(11, 2, '=IF(ABS(F9-SUM(F5:F8))<0.01,"ties","OUT BY "&TEXT(F9-SUM(F5:F8),"#,##0.0"))')
    ws.cell(12, 1, "Q3 growth over Q2").font = BOLD
    ws.cell(12, 2, "=D9/C9-1").number_format = "0.0%"
    ws.cell(13, 1, "Which is why the diff matters").font = BOLD
    ws.cell(13, 2, '="the typed total inflates Q3 growth to "&TEXT(D9/C9-1,"0.0%")'
                   '&" from "&TEXT(SUM(D5:D8)/C9-1,"0.0%")')
    return ws


def build() -> Workbook:
    wb = Workbook()
    traps(wb)
    circular(wb)
    hardcoded(wb)
    return wb
