"""M05 — Purchase approval workflow: nested logic, lookups and working days.

A deal-desk register. Every row is a purchase request that has to be routed to
an approval tier, assigned an approver by tier and region, tested for dual
approval, given an SLA, and checked for breach.

What it stresses: the logic vocabulary rather than the arithmetic one. Five-deep
nested IF, IFS, SWITCH, AND/OR combinations, a sorted approximate-match VLOOKUP
against the same thresholds the nested IF encodes (so the two must agree, cell by
cell), INDEX/MATCH into a two-dimensional approver grid, WORKDAY and NETWORKDAYS
with a holiday list, and text built by concatenation.

It also stresses laziness. `IFS` and `IF` must not evaluate the branches they do
not take — several rows here would raise #N/A or #DIV/0! if the untaken branch
were evaluated eagerly.

The as-of date is an input cell rather than TODAY(), which is what a generator
should write if the output is meant to be reproducible. TODAY() itself is
tested in M10.
"""

from __future__ import annotations

import random

from openpyxl import Workbook

from common import (
    BOLD, DATE, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY0, MONEY2, PCT,
    SUB_FILL, col, header_row, note, title, widths,
)

N = 160

DEPARTMENTS = ["Engineering", "Sales", "Marketing", "Finance", "Operations", "Legal", "People"]
REGIONS = ["EMEA", "Americas", "APAC"]
CATEGORIES = ["Software", "Hardware", "Consulting", "Travel", "Facilities", "Legal", "Marketing"]
PRIORITIES = ["Critical", "High", "Normal", "Low"]
STATUSES = ["Approved", "Pending", "In review", "Rejected"]
VENDOR_STATE = ["Approved vendor", "Unapproved vendor", "Sole source"]
TIERS = ["Auto-approve", "Manager", "Director", "VP", "CFO"]


def policy(wb: Workbook):
    ws = wb.active
    ws.title = "Policy"
    widths(ws, a=30, b=18, c=18, d=18, e=18)
    title(ws, "Approval policy")

    ws.cell(3, 1, "As-of date").font = BOLD
    ws.cell(3, 2, "=DATE(2026,7,20)").number_format = DATE
    ws.cell(4, 1, "Dual-approval amount threshold").font = BOLD
    ws.cell(4, 2, 50000).number_format = MONEY0
    ws.cell(4, 2).font = INPUT_FONT
    ws.cell(5, 1, "Dual-approval risk threshold").font = BOLD
    ws.cell(5, 2, 7).number_format = INT
    ws.cell(5, 2).font = INPUT_FONT
    ws.cell(6, 1, "Unapproved-vendor threshold").font = BOLD
    ws.cell(6, 2, 10000).number_format = MONEY0
    ws.cell(6, 2).font = INPUT_FONT

    # Sorted ascending — VLOOKUP's approximate match depends on it, and an
    # unsorted table here would be answered by Excel from its binary search's
    # probe order, which is a defensible thing for an engine to refuse.
    header_row(ws, 9, ["Amount from", "Tier", "SLA days", "Escalation hours"])
    bands = [
        (0, "Auto-approve", 0, 0),
        (1000, "Manager", 5, 48),
        (10000, "Director", 5, 24),
        (50000, "VP", 7, 12),
        (250000, "CFO", 10, 8),
    ]
    for i, (amount, tier, sla, hours) in enumerate(bands):
        r = 10 + i
        ws.cell(r, 1, amount).number_format = MONEY0
        ws.cell(r, 2, tier)
        ws.cell(r, 3, sla).number_format = INT
        ws.cell(r, 4, hours).number_format = INT

    ws.cell(16, 1, "Approver by tier and region").font = BOLD
    ws.cell(16, 1).fill = SUB_FILL
    header_row(ws, 17, ["Tier"] + REGIONS)
    approvers = {
        "Auto-approve": ["system", "system", "system"],
        "Manager": ["R. Okafor", "J. Whitfield", "M. Tanaka"],
        "Director": ["S. Lindqvist", "P. Ramírez", "H. Chandra"],
        "VP": ["A. Kowalski", "D. Brennan", "Y. Nakamura"],
        "CFO": ["E. Vandermeer", "E. Vandermeer", "E. Vandermeer"],
    }
    for i, tier in enumerate(TIERS):
        r = 18 + i
        ws.cell(r, 1, tier)
        for j, name in enumerate(approvers[tier]):
            ws.cell(r, 2 + j, name)

    ws.cell(24, 1, "Company holidays").font = BOLD
    ws.cell(24, 1).fill = SUB_FILL
    holidays = ["=DATE(2026,1,1)", "=DATE(2026,4,3)", "=DATE(2026,5,1)",
                "=DATE(2026,5,25)", "=DATE(2026,7,4)", "=DATE(2026,12,25)",
                "=DATE(2026,12,26)"]
    for i, h in enumerate(holidays):
        ws.cell(25 + i, 1, h).number_format = DATE

    ws.cell(33, 1, "Policy version").font = BOLD
    ws.cell(33, 2, '="v"&TEXT(2026,"0000")&"."&TEXT(7,"00")')
    return ws


def requests(wb: Workbook):
    ws = wb.create_sheet("Requests")
    title(ws, "Purchase request register")
    note(ws, 2, f"{N} requests. Columns L onward are derived by the policy on the Policy sheet.")
    ws.freeze_panes = "C5"

    cols = [
        ("Req ID", 9), ("Submitted", 12), ("Requester", 18), ("Department", 14),
        ("Region", 11), ("Category", 13), ("Amount", 13), ("Priority", 11),
        ("Risk", 7), ("Vendor status", 18), ("Status", 12),
        ("Tier (nested IF)", 15), ("Tier (lookup)", 14), ("Tiers agree", 13),
        ("Approver", 16), ("Dual approval", 14), ("SLA days", 10),
        ("Due date", 12), ("Working days open", 17), ("State", 14),
        ("Action required", 34), ("Risk band", 11), ("Reference", 22),
        ("Amount band", 13),
    ]
    header_row(ws, 4, [c[0] for c in cols])
    for i, (_label, w) in enumerate(cols):
        ws.column_dimensions[col(1 + i)].width = w

    rng = random.Random(4211)
    names = ["L. Persson", "K. Adeyemi", "T. Bianchi", "N. Haddad", "V. Sørensen",
             "C. Mwangi", "F. Delacroix", "G. Petrov", "I. Nakashima", "O. Bergström"]

    for k in range(N):
        r = 5 + k
        # A long tail of small requests with a few very large ones — the shape
        # that makes the tier bands worth having.
        magnitude = rng.choice([2, 2, 2, 3, 3, 3, 3, 4, 4, 5, 5, 6])
        amount = round(rng.uniform(1, 9.9) * 10 ** magnitude, 2)
        day = rng.randint(1, 190)
        ws.cell(r, 1, 100000 + k).number_format = "0"
        ws.cell(r, 2, f"=DATE(2026,1,1)+{day}").number_format = DATE
        ws.cell(r, 3, rng.choice(names))
        ws.cell(r, 4, rng.choice(DEPARTMENTS))
        ws.cell(r, 5, rng.choice(REGIONS))
        ws.cell(r, 6, rng.choice(CATEGORIES))
        ws.cell(r, 7, amount).number_format = MONEY2
        ws.cell(r, 8, rng.choice(PRIORITIES))
        ws.cell(r, 9, rng.randint(1, 10)).number_format = INT
        ws.cell(r, 10, rng.choices(VENDOR_STATE, weights=[7, 2, 1])[0])
        ws.cell(r, 11, rng.choices(STATUSES, weights=[5, 3, 2, 1])[0])

        # L: the policy as an analyst first writes it — nested IF.
        ws.cell(r, 12,
                f'=IF(G{r}<Policy!$A$11,"Auto-approve",'
                f'IF(G{r}<Policy!$A$12,"Manager",'
                f'IF(G{r}<Policy!$A$13,"Director",'
                f'IF(G{r}<Policy!$A$14,"VP","CFO"))))')
        # M: the same policy as a sorted approximate-match lookup.
        ws.cell(r, 13, f"=VLOOKUP(G{r},Policy!$A$10:$B$14,2,TRUE)")
        # N: and the assertion that the two encodings agree.
        ws.cell(r, 14, f'=IF(EXACT(L{r},M{r}),"yes","MISMATCH")')
        ws.cell(r, 15,
                f"=INDEX(Policy!$B$18:$D$22,MATCH(L{r},Policy!$A$18:$A$22,0),"
                f"MATCH(E{r},Policy!$B$17:$D$17,0))")
        ws.cell(r, 16,
                f'=IF(OR(AND(G{r}>=Policy!$B$4,I{r}>=Policy!$B$5),'
                f'F{r}="Legal",'
                f'AND(J{r}="Unapproved vendor",G{r}>=Policy!$B$6)),"Yes","No")')
        ws.cell(r, 17,
                f'=SWITCH(H{r},"Critical",1,"High",3,"Normal",'
                f"VLOOKUP(G{r},Policy!$A$10:$C$14,3,TRUE),10)").number_format = INT
        ws.cell(r, 18,
                f"=WORKDAY(B{r},MAX(1,Q{r}),Policy!$A$25:$A$31)").number_format = DATE
        ws.cell(r, 19,
                f"=NETWORKDAYS(B{r},Policy!$B$3,Policy!$A$25:$A$31)").number_format = INT
        ws.cell(r, 20,
                f'=IF(K{r}="Approved","closed",'
                f'IF(K{r}="Rejected","closed",'
                f'IF(Policy!$B$3>R{r},"BREACHED","in flight")))')
        # IFS, and it must be lazy: the first branch concatenates the approver,
        # which is an #N/A for any row whose tier is missing from the grid.
        ws.cell(r, 21,
                f'=IFS(T{r}="BREACHED","Escalate to "&O{r}&" ("'
                f'&TEXT(S{r}-Q{r},"0")&" days over)",'
                f'AND(P{r}="Yes",K{r}<>"Approved"),"Awaiting second approver",'
                f'K{r}="Rejected","No action — rejected",'
                f'K{r}="Approved","No action — approved",'
                f'TRUE,"On track")')
        ws.cell(r, 22, f'=IF(I{r}>=8,"High",IF(I{r}>=5,"Medium","Low"))')
        ws.cell(r, 23,
                f'=UPPER(LEFT(D{r},3))&"-"&TEXT(B{r},"yyyymm")&"-"&TEXT(A{r},"000000")')
        ws.cell(r, 24,
                f'=IF(G{r}>=1000000,"$1m+",TEXT(ROUNDDOWN(G{r}/10000,0)*10,"$#,##0")&"k")')
    return ws


def summary(wb: Workbook):
    ws = wb.create_sheet("Summary")
    widths(ws, a=32, b=14, c=14, d=14, e=14, f=14)
    title(ws, "Approval funnel")
    last = 4 + N
    amounts = f"Requests!$G$5:$G${last}"
    tiers = f"Requests!$L$5:$L${last}"
    status = f"Requests!$K$5:$K${last}"
    region = f"Requests!$E$5:$E${last}"
    dept = f"Requests!$D$5:$D${last}"
    state = f"Requests!$T$5:$T${last}"
    prio = f"Requests!$H$5:$H${last}"
    days = f"Requests!$S$5:$S${last}"
    risk = f"Requests!$I$5:$I${last}"

    header_row(ws, 4, ["Tier"] + STATUSES + ["Total value"])
    for i, tier in enumerate(TIERS):
        r = 5 + i
        ws.cell(r, 1, tier)
        for j, st in enumerate(STATUSES):
            ws.cell(r, 2 + j,
                    f'=COUNTIFS({tiers},$A{r},{status},{col(2 + j)}$4)').number_format = INT
        ws.cell(r, 6, f"=SUMIFS({amounts},{tiers},$A{r})").number_format = MONEY0
    ws.cell(10, 1, "Total").font = BOLD
    for j in range(len(STATUSES)):
        ws.cell(10, 2 + j, f"=SUM({col(2 + j)}5:{col(2 + j)}9)").number_format = INT
        ws.cell(10, 2 + j).font = BOLD
    ws.cell(10, 6, f"=SUM(F5:F9)").number_format = MONEY0
    ws.cell(10, 6).font = BOLD

    ws.cell(12, 1, "By region").font = BOLD
    ws.cell(12, 1).fill = SUB_FILL
    header_row(ws, 13, ["Region", "Requests", "Value", "Average", "Largest", "Breached"])
    for i, reg in enumerate(REGIONS):
        r = 14 + i
        ws.cell(r, 1, reg)
        ws.cell(r, 2, f"=COUNTIF({region},$A{r})").number_format = INT
        ws.cell(r, 3, f"=SUMIF({region},$A{r},{amounts})").number_format = MONEY0
        ws.cell(r, 4, f'=IFERROR(AVERAGEIFS({amounts},{region},$A{r}),"n/a")').number_format = MONEY0
        ws.cell(r, 5, f"=MAXIFS({amounts},{region},$A{r})").number_format = MONEY0
        ws.cell(r, 6, f'=COUNTIFS({region},$A{r},{state},"BREACHED")').number_format = INT

    ws.cell(19, 1, "Service levels").font = BOLD
    ws.cell(19, 1).fill = SUB_FILL
    stats = [
        ("Requests breaching SLA", f'=COUNTIF({state},"BREACHED")', INT),
        ("Breach rate", f'=COUNTIF({state},"BREACHED")/COUNTA({state})', PCT),
        ("Average working days open", f"=ROUND(AVERAGE({days}),1)", "0.0"),
        ("Longest open request",
         f"=INDEX(Requests!$A$5:$A${last},MATCH(MAX({days}),{days},0))", "0"),
        ("Critical requests still in flight",
         f'=COUNTIFS({prio},"Critical",{state},"in flight")', INT),
        ("Value awaiting approval", f'=SUMIFS({amounts},{status},"Pending")', MONEY0),
        ("Auto-approved share", f'=COUNTIF({tiers},"Auto-approve")/COUNTA({tiers})', PCT),
        ("High-risk approvals",
         f'=COUNTIFS({risk},">=8",{status},"Approved")', INT),
        ("Legal spend needing dual sign-off",
         f'=SUMIFS({amounts},Requests!$F$5:$F${last},"Legal",'
         f'Requests!$P$5:$P${last},"Yes")', MONEY0),
        ("Requests over $100k",
         f'=COUNTIF({amounts},">100000")', INT),
        ("Engineering share of spend",
         f'=SUMIF({dept},"Engineering",{amounts})/SUM({amounts})', PCT),
        ("Departments with any breach",
         f'=SUMPRODUCT((COUNTIFS({dept},Requests!$D$5:$D${last},{state},"BREACHED")>0)*1'
         f"/COUNTIF({dept},Requests!$D$5:$D${last}))", "0.0"),
    ]
    for i, (label, formula, fmt) in enumerate(stats):
        r = 20 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt

    ws.cell(34, 1, "Tier encodings agree on every row").font = BOLD
    ws.cell(34, 2,
            f'=IF(COUNTIF(Requests!$N$5:$N${last},"MISMATCH")=0,"consistent",'
            f'TEXT(COUNTIF(Requests!$N$5:$N${last},"MISMATCH"),"0")&" rows disagree")')
    ws.cell(35, 1, "Distinct approvers used").font = BOLD
    ws.cell(35, 2,
            f"=SUMPRODUCT(1/COUNTIF(Requests!$O$5:$O${last},"
            f"Requests!$O$5:$O${last}))").number_format = "0"
    ws.cell(36, 1, "Reference format sample").font = BOLD
    ws.cell(36, 2, "=Requests!W5")
    ws.cell(37, 1, "Policy applied").font = BOLD
    ws.cell(37, 2, '="Policy "&Policy!B33&" as of "&TEXT(Policy!B3,"d mmm yyyy")')
    return ws


def build() -> Workbook:
    wb = Workbook()
    policy(wb)
    requests(wb)
    summary(wb)
    return wb
