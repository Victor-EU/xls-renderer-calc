"""M09 — Inventory planning across 150 SKUs.

Demand history with a trend and a seasonal pattern, a statistical forecast, and
a per-SKU plan: safety stock at a target service level, reorder point, economic
order quantity, ABC classification and stockout risk.

What it stresses: the statistical and engineering end of the function library,
which is the part a finance-shaped test set never reaches. `NORM.S.INV` for the
service-level z-score, `SLOPE`/`INTERCEPT`/`FORECAST.LINEAR` for the trend,
`STDEV.S` for demand variability, `SQRT` in the EOQ, `PERCENTILE` and a running
cumulative share for the ABC cut.

This model was written without checking what the engine implements. If a
function here is refused, that is the finding — an operations model is a
perfectly ordinary thing to ask an agent for, and the failure mode to avoid is a
library shaped only like the tests that were written for it.
"""

from __future__ import annotations

import math
import random

from openpyxl import Workbook

from common import (
    BOLD, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY0, MONEY2, MONTH, NUM,
    PCT, SUB_FILL, col, header_row, note, title, widths,
)

SKUS = 150
MONTHS = 24
CATEGORIES = ["Fasteners", "Bearings", "Seals", "Castings", "Electronics", "Consumables"]


def demand(wb: Workbook):
    ws = wb.active
    ws.title = "Demand"
    widths(ws, a=16, b=12)
    title(ws, "Demand history by category")
    note(ws, 2, f"{MONTHS} months of shipped units, with a fitted trend and a six-month forecast.")

    header_row(ws, 4, ["Month", "Index"] + CATEGORIES)
    for i in range(MONTHS):
        ws.column_dimensions[col(3 + i)].width = 11

    rng = random.Random(31337)
    base = {c: rng.randint(1800, 9200) for c in CATEGORIES}
    trend = {c: rng.uniform(-0.004, 0.018) for c in CATEGORIES}
    for m in range(MONTHS):
        r = 5 + m
        ws.cell(r, 1, f"=EDATE(DATE(2024,7,1),{m})").number_format = MONTH
        ws.cell(r, 2, m + 1).number_format = INT
        for j, cat in enumerate(CATEGORIES):
            season = 1 + 0.18 * math.sin((m + j) * math.pi / 6)
            noise = rng.uniform(0.88, 1.12)
            units = round(base[cat] * (1 + trend[cat]) ** m * season * noise)
            ws.cell(r, 3 + j, units).number_format = INT

    last = 4 + MONTHS
    idx = f"$B$5:$B${last}"
    stats = [
        ("Mean", "=ROUND(AVERAGE({col}5:{col}%d),0)" % last, NUM),
        ("Std deviation", "=ROUND(STDEV.S({col}5:{col}%d),1)" % last, NUM),
        ("Coefficient of variation",
         "=STDEV.S({col}5:{col}%d)/AVERAGE({col}5:{col}%d)" % (last, last), PCT),
        ("Minimum", "=MIN({col}5:{col}%d)" % last, INT),
        ("Maximum", "=MAX({col}5:{col}%d)" % last, INT),
        ("Trend, units per month",
         "=SLOPE({col}5:{col}%d,%s)" % (last, idx), NUM),
        ("Intercept", "=INTERCEPT({col}5:{col}%d,%s)" % (last, idx), NUM),
        ("R squared", "=RSQ({col}5:{col}%d,%s)" % (last, idx), "0.000"),
        ("Last 3 months average", "=AVERAGE({col}%d:{col}%d)" % (last - 2, last), NUM),
        ("Momentum vs mean",
         "=AVERAGE({col}%d:{col}%d)/AVERAGE({col}5:{col}%d)-1" % (last - 2, last, last), PCT),
    ]
    for i, (label, tmpl, fmt) in enumerate(stats):
        r = last + 2 + i
        ws.cell(r, 1, label).font = BOLD
        for j in range(len(CATEGORIES)):
            c = col(3 + j)
            ws.cell(r, 3 + j, tmpl.replace("{col}", c)).number_format = fmt

    fbase = last + 2 + len(stats) + 1
    ws.cell(fbase, 1, "Six-month forecast").font = BOLD
    ws.cell(fbase, 1).fill = SUB_FILL
    header_row(ws, fbase + 1, ["Month", "Index"] + CATEGORIES)
    for h in range(6):
        r = fbase + 2 + h
        ws.cell(r, 1, f"=EDATE(DATE(2024,7,1),{MONTHS + h})").number_format = MONTH
        ws.cell(r, 2, MONTHS + h + 1).number_format = INT
        for j in range(len(CATEGORIES)):
            c = col(3 + j)
            ws.cell(r, 3 + j,
                    f"=ROUND(FORECAST.LINEAR($B{r},{c}5:{c}{last},{idx}),0)"
                    ).number_format = INT
    fc_last = fbase + 7
    ws.cell(fc_last + 2, 1, "Forecast total, next six months").font = BOLD
    ws.cell(fc_last + 2, 3,
            f"=SUM(C{fbase + 2}:{col(2 + len(CATEGORIES))}{fc_last})").number_format = INT
    ws.cell(fc_last + 3, 1, "Trailing six months actual").font = BOLD
    ws.cell(fc_last + 3, 3,
            f"=SUM(C{last - 5}:{col(2 + len(CATEGORIES))}{last})").number_format = INT
    ws.cell(fc_last + 4, 1, "Implied growth").font = BOLD
    ws.cell(fc_last + 4, 3,
            f"=C{fc_last + 2}/C{fc_last + 3}-1").number_format = PCT
    return ws, last


def items(wb: Workbook, demand_last: int):
    ws = wb.create_sheet("Items")
    title(ws, "SKU plan")
    note(ws, 2, f"{SKUS} SKUs. Safety stock is set from the target service level.")
    ws.freeze_panes = "C5"

    cols = [
        ("SKU", 11), ("Description", 22), ("Category", 14), ("Annual demand", 14),
        ("Unit cost", 11), ("Lead time (days)", 15), ("Demand std dev", 15),
        ("Service level", 12), ("Order cost", 11), ("Holding rate", 12),
        ("Annual value", 14), ("Value rank", 11), ("Cumulative %", 13),
        ("ABC class", 10), ("z-score", 10), ("Lead-time demand", 16),
        ("Safety stock", 13), ("Reorder point", 13), ("EOQ", 11),
        ("Orders per year", 14), ("Average inventory", 16),
        ("Inventory value", 15), ("Turns", 9), ("Days of cover", 13),
        ("Stockout risk", 13), ("Action", 26),
    ]
    header_row(ws, 4, [c[0] for c in cols])
    for i, (_label, w) in enumerate(cols):
        ws.column_dimensions[col(1 + i)].width = w

    rng = random.Random(5150)
    last = 4 + SKUS
    for k in range(SKUS):
        r = 5 + k
        cat = rng.choice(CATEGORIES)
        magnitude = rng.choice([2, 3, 3, 3, 4, 4, 5])
        annual = round(rng.uniform(1.2, 9.8) * 10 ** magnitude)
        cost = round(rng.uniform(0.8, 240.0), 2)
        ws.cell(r, 1, f'="SKU-"&TEXT({10000 + k * 7},"00000")')
        ws.cell(r, 2, f'=C{r}&" part "&TEXT({k + 1},"000")')
        ws.cell(r, 3, cat)
        ws.cell(r, 4, annual).number_format = INT
        ws.cell(r, 5, cost).number_format = MONEY2
        ws.cell(r, 6, rng.choice([7, 14, 21, 28, 45, 60, 90])).number_format = INT
        ws.cell(r, 7, round(annual / 12 * rng.uniform(0.12, 0.55), 1)).number_format = NUM
        ws.cell(r, 8, rng.choice([0.90, 0.95, 0.95, 0.975, 0.99])).number_format = PCT
        ws.cell(r, 9, rng.choice([45.0, 60.0, 85.0, 120.0])).number_format = MONEY0
        ws.cell(r, 10, 0.22).number_format = PCT

        ws.cell(r, 11, f"=D{r}*E{r}").number_format = MONEY0
        ws.cell(r, 12, f"=RANK.EQ(K{r},$K$5:$K${last},0)").number_format = INT
        # Cumulative share of annual value, ordered by rank — the ABC cut.
        ws.cell(r, 13,
                f'=SUMIF($L$5:$L${last},"<="&L{r},$K$5:$K${last})/SUM($K$5:$K${last})'
                ).number_format = PCT
        ws.cell(r, 14, f'=IF(M{r}<=0.8,"A",IF(M{r}<=0.95,"B","C"))')
        ws.cell(r, 15, f"=NORM.S.INV(H{r})").number_format = "0.000"
        ws.cell(r, 16, f"=D{r}/365*F{r}").number_format = NUM
        ws.cell(r, 17, f"=ROUND(O{r}*G{r}*SQRT(F{r}/30),0)").number_format = INT
        ws.cell(r, 18, f"=ROUND(P{r}+Q{r},0)").number_format = INT
        ws.cell(r, 19, f"=ROUND(SQRT(2*D{r}*I{r}/(E{r}*J{r})),0)").number_format = INT
        ws.cell(r, 20, f'=IFERROR(D{r}/S{r},"n/a")').number_format = NUM
        ws.cell(r, 21, f"=S{r}/2+Q{r}").number_format = NUM
        ws.cell(r, 22, f"=U{r}*E{r}").number_format = MONEY0
        ws.cell(r, 23, f'=IFERROR(D{r}/U{r},"n/a")').number_format = NUM
        ws.cell(r, 24, f'=IFERROR(365/W{r},"n/a")').number_format = NUM
        ws.cell(r, 25, f"=1-H{r}").number_format = PCT
        ws.cell(r, 26,
                f'=IF(AND(N{r}="A",X{r}>120),"Reduce cover — slow A item",'
                f'IF(AND(N{r}="A",H{r}<0.95),"Raise service level on an A item",'
                f'IF(AND(N{r}="C",V{r}>10000),"Overstocked C item",'
                f'IF(W{r}<2,"Low turns — review order quantity","No action"))))')
    return ws, last


def plan(wb: Workbook, items_last: int, demand_last: int):
    ws = wb.create_sheet("Plan")
    widths(ws, a=36, b=16, c=16, d=16, e=16, f=16)
    title(ws, "Inventory plan summary")

    I = "Items!"
    cls = f"{I}$N$5:$N${items_last}"
    val = f"{I}$V$5:$V${items_last}"
    ann = f"{I}$K$5:$K${items_last}"
    cat = f"{I}$C$5:$C${items_last}"
    turns = f"{I}$W$5:$W${items_last}"
    ss = f"{I}$Q$5:$Q${items_last}"
    eoq = f"{I}$S$5:$S${items_last}"
    lt = f"{I}$F$5:$F${items_last}"

    ws.cell(4, 1, "By ABC class").font = BOLD
    ws.cell(4, 1).fill = SUB_FILL
    header_row(ws, 5, ["Class", "SKUs", "Annual value", "Share", "Inventory value",
                       "Avg turns"])
    for i, c in enumerate(["A", "B", "C"]):
        r = 6 + i
        ws.cell(r, 1, c)
        ws.cell(r, 2, f"=COUNTIF({cls},$A{r})").number_format = INT
        ws.cell(r, 3, f"=SUMIF({cls},$A{r},{ann})").number_format = MONEY0
        ws.cell(r, 4, f"=C{r}/SUM({ann})").number_format = PCT
        ws.cell(r, 5, f"=SUMIF({cls},$A{r},{val})").number_format = MONEY0
        ws.cell(r, 6, f'=IFERROR(AVERAGEIF({cls},$A{r},{turns}),"n/a")').number_format = NUM
    ws.cell(9, 1, "Total").font = BOLD
    ws.cell(9, 2, "=SUM(B6:B8)").number_format = INT
    ws.cell(9, 3, "=SUM(C6:C8)").number_format = MONEY0
    ws.cell(9, 5, "=SUM(E6:E8)").number_format = MONEY0
    ws.cell(10, 1, "Class A share of value (Pareto check)").font = BOLD
    ws.cell(10, 3, "=C6/C9").number_format = PCT
    ws.cell(11, 1, "Pareto holds").font = BOLD
    ws.cell(11, 3, '=IF(C10>=0.7,"yes — A items dominate value","no")')

    ws.cell(13, 1, "By category").font = BOLD
    ws.cell(13, 1).fill = SUB_FILL
    header_row(ws, 14, ["Category", "SKUs", "Annual value", "Safety stock units",
                        "Average lead time", "Longest lead time"])
    for i, c in enumerate(CATEGORIES):
        r = 15 + i
        ws.cell(r, 1, c)
        ws.cell(r, 2, f"=COUNTIF({cat},$A{r})").number_format = INT
        ws.cell(r, 3, f"=SUMIF({cat},$A{r},{ann})").number_format = MONEY0
        ws.cell(r, 4, f"=SUMIF({cat},$A{r},{ss})").number_format = INT
        ws.cell(r, 5, f'=ROUND(AVERAGEIF({cat},$A{r},{lt}),1)').number_format = NUM
        ws.cell(r, 6, f"=MAXIFS({lt},{cat},$A{r})").number_format = INT

    base = 15 + len(CATEGORIES) + 1
    ws.cell(base, 1, "Portfolio").font = BOLD
    ws.cell(base, 1).fill = SUB_FILL
    stats = [
        ("SKUs planned", f"=COUNTA({cls})", INT),
        ("Total annual demand value", f"=SUM({ann})", MONEY0),
        ("Total inventory value", f"=SUM({val})", MONEY0),
        ("Inventory as % of annual value", f"=SUM({val})/SUM({ann})", PCT),
        ("Weighted average turns", f"=SUM({ann})/SUM({val})", NUM),
        ("Median turns", f"=MEDIAN({turns})", NUM),
        ("Slowest decile turns", f"=PERCENTILE.INC({turns},0.1)", NUM),
        ("Total safety stock units", f"=SUM({ss})", INT),
        ("Average EOQ", f"=ROUND(AVERAGE({eoq}),0)", INT),
        ("SKUs needing action",
         f'=COUNTIF({I}$Z$5:$Z${items_last},"<>No action")', INT),
        ("SKUs with turns below 2", f'=COUNTIF({turns},"<2")', INT),
        ("Longest lead time in the portfolio", f"=MAX({lt})", INT),
        ("Service level, value-weighted",
         f"=SUMPRODUCT({I}$H$5:$H${items_last},{ann})/SUM({ann})", PCT),
        ("Forecast six-month demand (all categories)",
         f"=Demand!C{demand_last + 22}", INT),
    ]
    for i, (label, formula, fmt) in enumerate(stats):
        r = base + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt

    top = base + len(stats) + 3
    ws.cell(top, 1, "Largest ten SKUs by inventory value").font = BOLD
    ws.cell(top, 1).fill = SUB_FILL
    header_row(ws, top + 1, ["Rank", "SKU", "Class", "Inventory value", "Turns"])
    for i in range(10):
        r = top + 2 + i
        ws.cell(r, 1, i + 1).number_format = INT
        ws.cell(r, 2,
                f"=INDEX({I}$A$5:$A${items_last},MATCH(LARGE({val},$A{r}),{val},0))")
        ws.cell(r, 3,
                f"=INDEX({cls},MATCH(LARGE({val},$A{r}),{val},0))")
        ws.cell(r, 4, f"=LARGE({val},$A{r})").number_format = MONEY0
        ws.cell(r, 5,
                f"=INDEX({turns},MATCH(LARGE({val},$A{r}),{val},0))").number_format = NUM
    ws.cell(top + 13, 1, "Top ten share of inventory value").font = BOLD
    ws.cell(top + 13, 4,
            f"=SUM(D{top + 2}:D{top + 11})/SUM({val})").number_format = PCT
    return ws


def build() -> Workbook:
    wb = Workbook()
    _, demand_last = demand(wb)
    _, items_last = items(wb, demand_last)
    plan(wb, items_last, demand_last)
    return wb
