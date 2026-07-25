"""M07 — SaaS cohort retention triangle.

Eighteen monthly cohorts down, eighteen months of age across, so half the grid
is genuinely empty — a cohort that started in month 14 has no month-17 data yet.

What it stresses: blank handling, which is the semantic corner where a
spreadsheet engine is most easily wrong. The upper triangle is `""` (the empty
string a formula returns), not a blank cell and not zero, and the three are
distinct to Excel. `AVERAGE` must skip them rather than count them as zero;
`COUNT` must not include them; `SUM` must tolerate them. Getting this subtly
wrong deflates every retention number on the sheet by a plausible-looking amount
— which is precisely the failure a reviewer cannot see.

The calendar-total row also exercises long formulas: each month's customer count
is an explicit sum of up to eighteen anti-diagonal cells, because that is what a
generator writes when the alternative is an array formula.
"""

from __future__ import annotations

from openpyxl import Workbook

from common import (
    BOLD, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY0, MONEY2, MONTH, NUM,
    PCT, SUB_FILL, col, header_row, note, title, widths,
)

COHORTS = 18
AGES = 18
FIRST = 4  # column D — age 0


def inputs(wb: Workbook):
    ws = wb.active
    ws.title = "Inputs"
    widths(ws, a=34, b=14)
    title(ws, "Cohort model inputs")

    rows = [
        ("First cohort month", None, MONTH),
        ("New customers, first cohort", 320, INT),
        ("New cohort growth, MoM", 0.060, PCT),
        ("Base monthly churn", 0.058, PCT),
        ("Churn improvement per month of age", 0.0035, PCT),
        ("Churn floor", 0.015, PCT),
        ("ARPU, month 0", 84.0, MONEY2),
        ("ARPU expansion, per month", 0.004, PCT),
        ("Gross margin", 0.780, PCT),
        ("Customer acquisition cost", 640.0, MONEY0),
        ("Monthly discount rate", 0.008, PCT),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value if value is not None else "=DATE(2026,1,1)")
        c.number_format = fmt
        if value is not None:
            c.font = INPUT_FONT

    ws.cell(15, 1, "Cohorts modelled").font = BOLD
    ws.cell(15, 2, COHORTS).number_format = INT
    ws.cell(16, 1, "Months of age tracked").font = BOLD
    ws.cell(16, 2, AGES).number_format = INT
    ws.cell(17, 1, "Last observable age index").font = BOLD
    ws.cell(17, 2, "=B16-1").number_format = INT

    ws.cell(19, 1, "Implied unit economics").font = BOLD
    ws.cell(19, 1).fill = SUB_FILL
    ws.cell(20, 1, "Gross margin per customer per month")
    ws.cell(20, 2, "=B9*B11").number_format = MONEY2
    ws.cell(21, 1, "Steady-state monthly churn")
    ws.cell(21, 2, "=MAX(B8,B6-B7*B16)").number_format = PCT
    ws.cell(22, 1, "Implied lifetime (months)")
    ws.cell(22, 2, "=1/B21").number_format = NUM
    ws.cell(23, 1, "Lifetime value")
    ws.cell(23, 2, "=B20/B21").number_format = MONEY0
    ws.cell(24, 1, "LTV / CAC")
    ws.cell(24, 2, "=B23/B12").number_format = "0.00"
    ws.cell(25, 1, "CAC payback (months)")
    ws.cell(25, 2, "=B12/B20").number_format = NUM
    ws.cell(26, 1, "Verdict")
    ws.cell(26, 2,
            '=IF(B24>=3,"healthy",IF(B24>=1.5,"acceptable","unsustainable"))')
    return ws


def triangle(wb: Workbook, name: str, title_text: str, cell_fn, fmt, header_note: str):
    ws = wb.create_sheet(name)
    widths(ws, a=13, b=8, c=12)
    for n in range(AGES):
        ws.column_dimensions[col(FIRST + n)].width = 10
    title(ws, title_text)
    note(ws, 2, header_note)
    ws.freeze_panes = "D5"

    header_row(ws, 4, ["Cohort", "Index", "Size"] + [f"M{n}" for n in range(AGES)])
    for i in range(COHORTS):
        r = 5 + i
        ws.cell(r, 1, f"=EDATE(Inputs!$B$3,{i})").number_format = MONTH
        ws.cell(r, 2, i).number_format = INT
        ws.cell(r, 3,
                f"=ROUND(Inputs!$B$4*(1+Inputs!$B$5)^{i},0)").number_format = INT
        for n in range(AGES):
            cell = ws.cell(r, FIRST + n, cell_fn(i, n, r, col(FIRST + n), col(FIRST + n - 1)))
            cell.number_format = fmt
    return ws


def cohorts(wb: Workbook):
    def fn(i, n, r, c, prev):
        # Above the diagonal there is no data yet — an empty string, which is
        # what a real cohort sheet shows and what downstream stats must skip.
        guard = f"$B{r}+{n}>Inputs!$B$17"
        if n == 0:
            return f'=IF({guard},"",$C{r})'
        churn = "MAX(Inputs!$B$8,Inputs!$B$6-Inputs!$B$7*%d)" % n
        return f'=IF({guard},"",ROUND({prev}{r}*(1-{churn}),1))'

    return triangle(wb, "Cohorts", "Customers retained by cohort age", fn, NUM,
                    "Blank cells above the diagonal are months that have not happened yet.")


def revenue(wb: Workbook):
    def fn(i, n, r, c, prev):
        return (f'=IF(Cohorts!{c}{r}="","",'
                f'ROUND(Cohorts!{c}{r}*Inputs!$B$9*(1+Inputs!$B$10)^{n},0))')

    return triangle(wb, "Revenue", "Monthly recurring revenue by cohort age", fn, MONEY0,
                    "Mirrors the retention triangle; ARPU expands with tenure.")


def metrics(wb: Workbook):
    ws = wb.create_sheet("Metrics")
    widths(ws, a=34, b=14, c=14, d=14)
    title(ws, "Retention and unit economics")

    last = 4 + COHORTS
    sizes = f"Cohorts!$C$5:$C${last}"

    ws.cell(4, 1, "Average retention curve").font = BOLD
    ws.cell(4, 1).fill = SUB_FILL
    header_row(ws, 5, ["Age", "Customers", "Base cohorts", "Retention", "Monthly churn"])
    for n in range(AGES):
        r = 6 + n
        band = f"Cohorts!{col(FIRST + n)}$5:{col(FIRST + n)}${last}"
        ws.cell(r, 1, n).number_format = INT
        ws.cell(r, 2, f"=SUM({band})").number_format = NUM
        # Only cohorts old enough to have this month should be in the denominator.
        ws.cell(r, 3, f'=SUMIF({band},">0",{sizes})').number_format = NUM
        ws.cell(r, 4, f'=IFERROR(B{r}/C{r},"")').number_format = PCT
        ws.cell(r, 5,
                ('=""' if n == 0 else f'=IFERROR(1-D{r}/D{r - 1},"")')).number_format = PCT

    base = 6 + AGES + 1
    ws.cell(base, 1, "Retention checkpoints").font = BOLD
    ws.cell(base, 1).fill = SUB_FILL
    checkpoints = [
        ("Month 1 retention", "=D7", PCT),
        ("Month 3 retention", "=D9", PCT),
        ("Month 6 retention", "=D12", PCT),
        ("Month 12 retention", "=D18", PCT),
        ("Average churn, months 1-6", "=AVERAGE(E7:E12)", PCT),
        ("Average churn, months 7-12", "=AVERAGE(E13:E18)", PCT),
        ("Churn is improving with tenure",
         '=IF(B%d<B%d,"yes","no")' % (base + 5, base + 4), "General"),
    ]
    for i, (label, formula, fmt) in enumerate(checkpoints):
        r = base + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt

    # Calendar totals are anti-diagonals: in calendar month m, cohort i is at
    # age m-i. Written out term by term, which is what a generator produces when
    # the alternative needs an array formula.
    base2 = base + len(checkpoints) + 3
    ws.cell(base2, 1, "Calendar month totals").font = BOLD
    ws.cell(base2, 1).fill = SUB_FILL
    header_row(ws, base2 + 1, ["Month", "Customers", "MRR", "New", "Net adds"])
    for m in range(AGES):
        r = base2 + 2 + m
        terms_c = "+".join(f"Cohorts!{col(FIRST + m - i)}{5 + i}" for i in range(m + 1))
        terms_r = "+".join(f"Revenue!{col(FIRST + m - i)}{5 + i}" for i in range(m + 1))
        ws.cell(r, 1, f"=EDATE(Inputs!$B$3,{m})").number_format = MONTH
        ws.cell(r, 2, f"={terms_c}").number_format = NUM
        ws.cell(r, 3, f"={terms_r}").number_format = MONEY0
        ws.cell(r, 4, f"=Cohorts!$C${5 + m}").number_format = INT
        ws.cell(r, 5, ("=B%d" % r if m == 0 else f"=B{r}-B{r - 1}")).number_format = NUM

    base3 = base2 + AGES + 3
    cust_col = f"$B${base2 + 2}:$B${base2 + 1 + AGES}"
    mrr_col = f"$C${base2 + 2}:$C${base2 + 1 + AGES}"
    ws.cell(base3, 1, "Business summary").font = BOLD
    ws.cell(base3, 1).fill = SUB_FILL
    summary = [
        ("Customers at latest month", f"=INDEX({cust_col},{AGES})", NUM),
        ("MRR at latest month", f"=INDEX({mrr_col},{AGES})", MONEY0),
        ("Peak MRR", f"=MAX({mrr_col})", MONEY0),
        ("MRR growth, last month", f"=INDEX({mrr_col},{AGES})/INDEX({mrr_col},{AGES - 1})-1", PCT),
        ("Cumulative customers acquired", f"=SUM({sizes})", NUM),
        ("Customers retained", f"=INDEX({cust_col},{AGES})", NUM),
        ("Lifetime retention rate",
         f"=INDEX({cust_col},{AGES})/SUM({sizes})", PCT),
        ("Total MRR booked", f"=SUM({mrr_col})", MONEY0),
        ("ARPU, blended latest",
         f"=INDEX({mrr_col},{AGES})/INDEX({cust_col},{AGES})", MONEY2),
        ("Contribution margin, latest month",
         f"=INDEX({mrr_col},{AGES})*Inputs!$B$11", MONEY0),
        ("Acquisition spend to date", f"=SUM({sizes})*Inputs!$B$12", MONEY0),
        ("Payback status",
         f'=IF(SUM({mrr_col})*Inputs!$B$11>SUM({sizes})*Inputs!$B$12,'
         f'"cohorts have paid back","not yet paid back")', "General"),
        ("NPV of latest cohort's remaining margin",
         f"=Inputs!$B$20*INDEX({cust_col},{AGES})/(Inputs!$B$13+Inputs!$B$21)", MONEY0),
    ]
    for i, (label, formula, fmt) in enumerate(summary):
        r = base3 + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt

    base4 = base3 + len(summary) + 3
    ws.cell(base4, 1, "Blank-handling assertions").font = BOLD
    ws.cell(base4, 1).fill = SUB_FILL
    grid = f"Cohorts!$D$5:${col(FIRST + AGES - 1)}${last}"
    assertions = [
        ("Cells in the triangle", f"={COHORTS}*{AGES}", INT),
        ("Numeric cells (COUNT)", f"=COUNT({grid})", INT),
        ("Non-empty cells (COUNTA)", f"=COUNTA({grid})", INT),
        ("Empty-string cells", f'=COUNTIF({grid},"")', INT),
        ("COUNTA counts the empty strings too",
         f'=IF(B{base4 + 3}=B{base4 + 1},"yes — COUNTA sees \'\' as content","no")', "General"),
        ("Filled cells match the triangle shape",
         f'=IF(B{base4 + 2}={COHORTS}*{AGES}-{COHORTS}*({COHORTS}-1)/2,'
         f'"triangle is the expected shape","UNEXPECTED")', "General"),
        ("Average ignores the empty strings",
         f'=IF(ABS(AVERAGE({grid})-SUM({grid})/COUNT({grid}))<0.0001,"yes","NO")', "General"),
    ]
    for i, (label, formula, fmt) in enumerate(assertions):
        r = base4 + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt
    return ws


def build() -> Workbook:
    wb = Workbook()
    inputs(wb)
    cohorts(wb)
    revenue(wb)
    metrics(wb)
    return wb
