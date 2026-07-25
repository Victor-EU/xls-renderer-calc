"""M06 — Two thousand transactions with a criteria-driven analysis layer.

A flat sales ledger of the kind that comes out of a CRM export, with ten derived
columns per row, and an analysis sheet that slices it the way a spreadsheet
does: SUMIFS, COUNTIFS, AVERAGEIFS, MAXIFS, wildcards, and date-range criteria
built by concatenation.

What it stresses:

  scale        ~20,000 formula cells in one sheet, all reading a 2,000-row range
  criteria     `">="&DATE(...)` is a string built at evaluation time; `"AX-*"`
               is a wildcard; `">"&$B$4` compares against a cell. Each is parsed
               differently and each is a place to be quietly wrong
  agreement    several totals are computed twice, once with SUMIFS and once with
               SUMPRODUCT. They must agree — and they are the kind of pair a
               criteria bug breaks asymmetrically
"""

from __future__ import annotations

import random

from openpyxl import Workbook

from common import (
    BOLD, DATE, HDR_FILL, HDR_FONT, INT, MONEY0, MONEY2, NUM, PCT, PCT2,
    SUB_FILL, col, header_row, note, title, widths,
)

N = 2000

REGIONS = ["North America", "Northern Europe", "Southern Europe", "APAC", "LATAM"]
COUNTRY = {
    "North America": ["United States", "Canada"],
    "Northern Europe": ["United Kingdom", "Sweden", "Netherlands"],
    "Southern Europe": ["Spain", "Italy", "Portugal"],
    "APAC": ["Japan", "Australia", "Singapore"],
    "LATAM": ["Brazil", "Mexico", "Chile"],
}
CHANNELS = ["Direct", "Partner", "Self-serve", "Marketplace"]
REPS = ["A. Ferreira", "B. Lindgren", "C. Osei", "D. Marchetti", "E. Novak",
        "F. Yamamoto", "G. Rousseau", "H. Almeida", "I. Kaur", "J. Sørensen",
        "K. Mbeki", "L. Fitzgerald"]

SKUS = [
    ("AX-1000", "Platform", 420.0), ("AX-1100", "Platform", 480.0),
    ("AX-1200", "Platform", 610.0), ("AX-2000", "Platform", 890.0),
    ("BN-3000", "Analytics", 240.0), ("BN-3100", "Analytics", 310.0),
    ("BN-3200", "Analytics", 375.0), ("CD-4000", "Connectors", 95.0),
    ("CD-4100", "Connectors", 120.0), ("CD-4200", "Connectors", 145.0),
    ("DE-5000", "Services", 1200.0), ("DE-5100", "Services", 1650.0),
    ("EF-6000", "Support", 60.0), ("EF-6100", "Support", 85.0),
    ("EF-6200", "Support", 110.0), ("FG-7000", "Training", 300.0),
]


def lookups(wb: Workbook):
    ws = wb.active
    ws.title = "Lookups"
    widths(ws, a=14, b=16, c=14, d=14, e=14)
    title(ws, "Product master")
    header_row(ws, 4, ["SKU", "Family", "Unit cost", "List price", "Margin %"])
    for i, (sku, family, cost) in enumerate(SKUS):
        r = 5 + i
        ws.cell(r, 1, sku)
        ws.cell(r, 2, family)
        ws.cell(r, 3, cost).number_format = MONEY2
        ws.cell(r, 4, f"=ROUND(C{r}/(1-0.62),2)").number_format = MONEY2
        ws.cell(r, 5, f"=1-C{r}/D{r}").number_format = PCT
    last = 4 + len(SKUS)
    ws.cell(last + 2, 1, "Families").font = BOLD
    families = sorted({f for _s, f, _c in SKUS})
    for i, f in enumerate(families):
        ws.cell(last + 3 + i, 1, f)
        ws.cell(last + 3 + i, 2, f'=COUNTIF($B$5:$B${last},A{last + 3 + i})').number_format = INT
    return ws, families


def transactions(wb: Workbook):
    ws = wb.create_sheet("Transactions")
    title(ws, "Order ledger — FY2026 year to date")
    note(ws, 2, f"{N:,} orders. Columns K onward are derived.")
    ws.freeze_panes = "C5"

    cols = [
        ("Order date", 12), ("Order ID", 12), ("Region", 17), ("Country", 15),
        ("Sales rep", 15), ("Channel", 13), ("SKU", 11), ("Units", 8),
        ("Unit price", 11), ("Discount", 10),
        ("Gross revenue", 14), ("Net revenue", 13), ("Family", 13),
        ("Unit cost", 11), ("COGS", 12), ("Margin", 12), ("Margin %", 10),
        ("Quarter", 9), ("Month", 10), ("Deal band", 13),
    ]
    header_row(ws, 4, [c[0] for c in cols])
    for i, (_label, w) in enumerate(cols):
        ws.column_dimensions[col(1 + i)].width = w

    rng = random.Random(90210)
    sku_last = 4 + len(SKUS)
    for k in range(N):
        r = 5 + k
        region = rng.choice(REGIONS)
        sku, _family, _cost = rng.choice(SKUS)
        # Seasonal skew: more orders late in the period, so the date-range
        # criteria on the analysis sheet actually separate different numbers.
        day = int(rng.triangular(0, 200, 150))
        units = rng.choice([1, 1, 2, 2, 3, 5, 8, 10, 12, 25, 40, 60])
        price = round(next(p for s, _f, p in SKUS if s == sku) * rng.uniform(1.5, 2.9), 2)
        discount = rng.choice([0, 0, 0, 0.05, 0.05, 0.10, 0.125, 0.15, 0.20, 0.30])

        ws.cell(r, 1, f"=DATE(2026,1,1)+{day}").number_format = DATE
        ws.cell(r, 2, f'="ORD-"&TEXT({500000 + k},"000000")')
        ws.cell(r, 3, region)
        ws.cell(r, 4, rng.choice(COUNTRY[region]))
        ws.cell(r, 5, rng.choice(REPS))
        ws.cell(r, 6, rng.choices(CHANNELS, weights=[5, 3, 3, 1])[0])
        ws.cell(r, 7, sku)
        ws.cell(r, 8, units).number_format = INT
        ws.cell(r, 9, price).number_format = MONEY2
        ws.cell(r, 10, discount).number_format = PCT

        ws.cell(r, 11, f"=H{r}*I{r}").number_format = MONEY2
        ws.cell(r, 12, f"=ROUND(K{r}*(1-J{r}),2)").number_format = MONEY2
        ws.cell(r, 13, f"=VLOOKUP($G{r},Lookups!$A$5:$C${sku_last},2,FALSE)")
        ws.cell(r, 14, f"=VLOOKUP($G{r},Lookups!$A$5:$C${sku_last},3,FALSE)").number_format = MONEY2
        ws.cell(r, 15, f"=ROUND(H{r}*N{r},2)").number_format = MONEY2
        ws.cell(r, 16, f"=L{r}-O{r}").number_format = MONEY2
        ws.cell(r, 17, f'=IFERROR(P{r}/L{r},"")').number_format = PCT
        ws.cell(r, 18, f'="Q"&ROUNDUP(MONTH(A{r})/3,0)')
        ws.cell(r, 19, f'=TEXT(A{r},"yyyy-mm")')
        ws.cell(r, 20,
                f'=IF(L{r}>=50000,"Enterprise",IF(L{r}>=10000,"Mid-market",'
                f'IF(L{r}>=1000,"SMB","Micro")))')
    return ws


def analysis(wb: Workbook, families):
    ws = wb.create_sheet("Analysis")
    widths(ws, a=30, b=15, c=15, d=15, e=15, f=15, g=15, h=15)
    title(ws, "Revenue analysis")
    last = 4 + N
    T = f"Transactions!"
    date_r = f"{T}$A$5:$A${last}"
    region_r = f"{T}$C$5:$C${last}"
    country_r = f"{T}$D$5:$D${last}"
    rep_r = f"{T}$E$5:$E${last}"
    chan_r = f"{T}$F$5:$F${last}"
    sku_r = f"{T}$G$5:$G${last}"
    units_r = f"{T}$H$5:$H${last}"
    disc_r = f"{T}$J$5:$J${last}"
    net_r = f"{T}$L$5:$L${last}"
    fam_r = f"{T}$M$5:$M${last}"
    margin_r = f"{T}$P$5:$P${last}"
    qtr_r = f"{T}$R$5:$R${last}"
    band_r = f"{T}$T$5:$T${last}"

    # --- region x family -----------------------------------------------------
    ws.cell(4, 1, "Net revenue by region and family").font = BOLD
    ws.cell(4, 1).fill = SUB_FILL
    header_row(ws, 5, ["Region"] + families + ["Total"])
    for i, region in enumerate(REGIONS):
        r = 6 + i
        ws.cell(r, 1, region)
        for j, fam in enumerate(families):
            ws.cell(r, 2 + j,
                    f"=SUMIFS({net_r},{region_r},$A{r},{fam_r},{col(2 + j)}$5)"
                    ).number_format = MONEY0
        ws.cell(r, 2 + len(families),
                f"=SUM(B{r}:{col(1 + len(families))}{r})").number_format = MONEY0
        ws.cell(r, 2 + len(families)).font = BOLD
    tot = 6 + len(REGIONS)
    ws.cell(tot, 1, "Total").font = BOLD
    for j in range(len(families) + 1):
        c = col(2 + j)
        ws.cell(tot, 2 + j, f"=SUM({c}6:{c}{tot - 1})").number_format = MONEY0
        ws.cell(tot, 2 + j).font = BOLD
    ws.cell(tot + 1, 1, "Cross-check against SUM of the column").font = BOLD
    ws.cell(tot + 1, 2,
            f'=IF(ABS({col(2 + len(families))}{tot}-SUM({net_r}))<0.5,"ties",'
            f'"OUT BY "&TEXT({col(2 + len(families))}{tot}-SUM({net_r}),"#,##0.00"))')

    # --- quarter x channel ---------------------------------------------------
    base = tot + 3
    ws.cell(base, 1, "Margin by quarter and channel").font = BOLD
    ws.cell(base, 1).fill = SUB_FILL
    header_row(ws, base + 1, ["Quarter"] + CHANNELS + ["Total", "Margin %"])
    for i in range(4):
        r = base + 2 + i
        ws.cell(r, 1, f'="Q"&{i + 1}')
        for j, chan in enumerate(CHANNELS):
            ws.cell(r, 2 + j,
                    f"=SUMIFS({margin_r},{qtr_r},$A{r},{chan_r},{col(2 + j)}${base + 1})"
                    ).number_format = MONEY0
        ws.cell(r, 6, f"=SUM(B{r}:E{r})").number_format = MONEY0
        ws.cell(r, 7,
                f'=IFERROR(F{r}/SUMIFS({net_r},{qtr_r},$A{r}),"n/a")').number_format = PCT

    # --- criteria forms ------------------------------------------------------
    base2 = base + 7
    ws.cell(base2, 1, "Criteria forms").font = BOLD
    ws.cell(base2, 1).fill = SUB_FILL
    criteria = [
        ("Revenue on or after 1 Jul", f'=SUMIFS({net_r},{date_r},">="&DATE(2026,7,1))', MONEY0),
        ("Revenue in H1", f'=SUMIFS({net_r},{date_r},"<"&DATE(2026,7,1))', MONEY0),
        ("H1 plus H2 equals total",
         f'=IF(ABS(B{base2 + 1}+B{base2 + 2}-SUM({net_r}))<0.5,"ties","GAP")', "General"),
        ("Orders on platform SKUs (wildcard)", f'=COUNTIF({sku_r},"AX-*")', INT),
        ("Platform revenue (wildcard)", f'=SUMIF({sku_r},"AX-*",{net_r})', MONEY0),
        ("Platform revenue (family match)",
         f'=SUMIF({fam_r},"Platform",{net_r})', MONEY0),
        ("Wildcard and family agree",
         f'=IF(ABS(B{base2 + 5}-B{base2 + 6})<0.5,"ties","differ — SKU prefix is not the family")',
         "General"),
        ("Orders above the mean", f'=COUNTIF({net_r},">"&AVERAGE({net_r}))', INT),
        ("Discounted order count", f'=COUNTIF({disc_r},">0")', INT),
        ("Undiscounted revenue", f'=SUMIF({disc_r},0,{net_r})', MONEY0),
        ("Deep-discount revenue (>=20%)", f'=SUMIF({disc_r},">=0.2",{net_r})', MONEY0),
        ("Enterprise deals", f'=COUNTIF({band_r},"Enterprise")', INT),
        ("Single-letter wildcard test", f'=COUNTIF({sku_r},"AX-1?00")', INT),
        ("Countries containing 'land'", f'=COUNTIF({country_r},"*land*")', INT),
    ]
    for i, (label, formula, fmt) in enumerate(criteria):
        r = base2 + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt

    # --- rep league table ----------------------------------------------------
    base3 = base2 + len(criteria) + 3
    ws.cell(base3, 1, "Sales rep league table").font = BOLD
    ws.cell(base3, 1).fill = SUB_FILL
    header_row(ws, base3 + 1,
               ["Rep", "Net revenue", "Orders", "Average order", "Margin %",
                "Rank", "Share"])
    for i, rep in enumerate(REPS):
        r = base3 + 2 + i
        ws.cell(r, 1, rep)
        ws.cell(r, 2, f"=SUMIF({rep_r},$A{r},{net_r})").number_format = MONEY0
        ws.cell(r, 3, f"=COUNTIF({rep_r},$A{r})").number_format = INT
        ws.cell(r, 4, f'=IFERROR(B{r}/C{r},"n/a")').number_format = MONEY0
        ws.cell(r, 5,
                f'=IFERROR(SUMIF({rep_r},$A{r},{margin_r})/B{r},"n/a")').number_format = PCT
        ws.cell(r, 6,
                f"=RANK.EQ(B{r},$B${base3 + 2}:$B${base3 + 1 + len(REPS)},0)").number_format = INT
        ws.cell(r, 7, f"=B{r}/SUM({net_r})").number_format = PCT2
    rep_last = base3 + 1 + len(REPS)
    rep_rev = f"$B${base3 + 2}:$B${rep_last}"
    rep_name = f"$A${base3 + 2}:$A${rep_last}"

    base4 = rep_last + 2
    tops = [
        ("Top rep", f"=INDEX({rep_name},MATCH(MAX({rep_rev}),{rep_rev},0))", "General"),
        ("Top rep revenue", f"=MAX({rep_rev})", MONEY0),
        ("Second place", f"=INDEX({rep_name},MATCH(LARGE({rep_rev},2),{rep_rev},0))", "General"),
        ("Third place", f"=INDEX({rep_name},MATCH(LARGE({rep_rev},3),{rep_rev},0))", "General"),
        ("Top three share", f"=SUM(LARGE({rep_rev},1),LARGE({rep_rev},2),"
                            f"LARGE({rep_rev},3))/SUM({rep_rev})", PCT),
        ("Bottom rep", f"=INDEX({rep_name},MATCH(MIN({rep_rev}),{rep_rev},0))", "General"),
        ("Reps above average", f"=COUNTIF({rep_rev},\">\"&AVERAGE({rep_rev}))", INT),
        ("Revenue concentration (max / total)", f"=MAX({rep_rev})/SUM({rep_rev})", PCT),
    ]
    for i, (label, formula, fmt) in enumerate(tops):
        r = base4 + i
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 2, formula).number_format = fmt

    # --- totals computed two ways -------------------------------------------
    base5 = base4 + len(tops) + 2
    ws.cell(base5, 1, "Totals, computed two ways").font = BOLD
    ws.cell(base5, 1).fill = SUB_FILL
    pairs = [
        ("Direct-channel revenue — SUMIFS", f'=SUMIFS({net_r},{chan_r},"Direct")'),
        ("Direct-channel revenue — SUMPRODUCT",
         f'=SUMPRODUCT(({chan_r}="Direct")*{net_r})'),
        ("APAC units — SUMIFS", f'=SUMIFS({units_r},{region_r},"APAC")'),
        ("APAC units — SUMPRODUCT", f'=SUMPRODUCT(({region_r}="APAC")*{units_r})'),
        ("Q4 partner revenue — SUMIFS",
         f'=SUMIFS({net_r},{qtr_r},"Q4",{chan_r},"Partner")'),
        ("Q4 partner revenue — SUMPRODUCT",
         f'=SUMPRODUCT(({qtr_r}="Q4")*({chan_r}="Partner")*{net_r})'),
    ]
    for i, (label, formula) in enumerate(pairs):
        r = base5 + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = MONEY0
    for i in range(3):
        r = base5 + 1 + i * 2
        ws.cell(r, 3,
                f'=IF(ABS(B{r}-B{r + 1})<0.5,"agree","DIFFER BY "&TEXT(B{r}-B{r + 1},"#,##0.00"))')

    # --- distribution --------------------------------------------------------
    base6 = base5 + len(pairs) + 3
    ws.cell(base6, 1, "Order size distribution").font = BOLD
    ws.cell(base6, 1).fill = SUB_FILL
    dist = [
        ("Orders", f"=COUNT({net_r})", INT),
        ("Total net revenue", f"=SUM({net_r})", MONEY0),
        ("Mean order", f"=AVERAGE({net_r})", MONEY0),
        ("Median order", f"=MEDIAN({net_r})", MONEY0),
        ("90th percentile", f"=PERCENTILE.INC({net_r},0.9)", MONEY0),
        ("Standard deviation", f"=STDEV.S({net_r})", MONEY0),
        ("Largest order", f"=MAX({net_r})", MONEY0),
        ("Smallest order", f"=MIN({net_r})", MONEY0),
        ("Blended margin", f"=SUM({margin_r})/SUM({net_r})", PCT),
        ("Average discount, weighted",
         f"=SUMPRODUCT({disc_r},{net_r})/SUM({net_r})", PCT2),
        ("Discount vs order size, correlation",
         f"=CORREL({disc_r},{net_r})", "0.000"),
        ("Units per order", f"=SUM({units_r})/COUNT({net_r})", NUM),
    ]
    for i, (label, formula, fmt) in enumerate(dist):
        r = base6 + 1 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, formula).number_format = fmt
    return ws


def build() -> Workbook:
    wb = Workbook()
    _, families = lookups(wb)
    transactions(wb)
    analysis(wb, families)
    return wb
