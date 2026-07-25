"""M01 — Annual operating budget with variance analysis.

The most common thing a finance user asks an agent for. Twelve months across,
line items down, a plan sheet driven by assumptions, actuals beside it, and a
variance sheet that flags what needs attention.

What it stresses: cross-sheet references with mixed anchoring, a running YTD
total (`SUM($B$7:F7)`) which is the dependency shape that made the naive graph
representation blow up, sign conventions where costs are negative, IFERROR on
margins, and text flags built from comparisons.
"""

from __future__ import annotations

import random

from openpyxl import Workbook

from common import (
    BOLD, DATE, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY, MONTH, PCT,
    SUB_FILL, TOP_BORDER, TOTAL_BORDER, col, header_row, note, title, widths,
)

MONTHS = 12
FIRST = 2  # column B

# Line items, as (row, label, kind) — kind drives the variance sign convention.
LINES = [
    (5, "Revenue — Product", "income"),
    (6, "Revenue — Services", "income"),
    (7, "Total revenue", "income"),
    (9, "COGS — Product", "cost"),
    (10, "COGS — Services", "cost"),
    (11, "Total COGS", "cost"),
    (12, "Gross profit", "income"),
    (15, "Salaries & wages", "cost"),
    (16, "Benefits & payroll tax", "cost"),
    (17, "Marketing", "cost"),
    (18, "Facilities", "cost"),
    (19, "Software & IT", "cost"),
    (20, "Travel & entertainment", "cost"),
    (21, "Professional fees", "cost"),
    (22, "Total operating expenses", "cost"),
    (24, "EBITDA", "income"),
]


def assumptions(wb: Workbook):
    ws = wb.active
    ws.title = "Assumptions"
    widths(ws, a=30, b=14, c=14)
    title(ws, "Operating budget — assumptions")
    note(ws, 2, "Blue cells are inputs. Everything else is calculated.")

    rows = [
        ("Product revenue, month 1", 850.0, MONEY),
        ("Services revenue, month 1", 320.0, MONEY),
        ("Product revenue growth, MoM", 0.021, PCT),
        ("Services revenue growth, MoM", 0.015, PCT),
        ("Product COGS, % of revenue", 0.28, PCT),
        ("Services COGS, % of revenue", 0.55, PCT),
        ("Marketing, % of revenue", 0.14, PCT),
        ("Travel, % of revenue", 0.018, PCT),
        ("Facilities, per month", 68.0, MONEY),
        ("Software & IT, per month", 41.0, MONEY),
        ("Professional fees, per month", 25.0, MONEY),
        ("Benefits, % of salary", 0.22, PCT),
        ("Average fully-loaded salary", 142000.0, INT),
        ("Fiscal year start", "2026-01-01", DATE),
        ("Variance threshold, absolute", 25.0, MONEY),
        ("Variance threshold, percent", 0.05, PCT),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value)
        c.number_format = fmt
        c.font = INPUT_FONT

    # openpyxl writes a date string as text; make it a real serial via DATE().
    ws["B16"] = "=DATE(2026,1,1)"
    ws["B16"].number_format = DATE

    header_row(ws, 20, ["Month", "Seasonality", "Headcount"])
    seasonality = [0.92, 0.94, 1.03, 1.01, 1.05, 1.08, 0.96, 0.93, 1.07, 1.09, 1.02, 0.90]
    headcount = [64, 64, 67, 69, 69, 72, 74, 74, 77, 79, 81, 82]
    for i in range(MONTHS):
        ws.cell(21 + i, 1, f"=TEXT(EDATE($B$16,{i}),\"mmm-yy\")")
        c = ws.cell(21 + i, 2, seasonality[i])
        c.number_format = "0.00"
        c.font = INPUT_FONT
        c = ws.cell(21 + i, 3, headcount[i])
        c.number_format = INT
        c.font = INPUT_FONT

    ws.cell(34, 1, "Average headcount").font = BOLD
    ws.cell(34, 3, "=ROUND(AVERAGE(C21:C32),1)").number_format = "0.0"
    ws.cell(35, 1, "Seasonality check (should be ~12.0)").font = BOLD
    ws.cell(35, 2, "=SUM(B21:B32)").number_format = "0.00"
    return ws


def budget(wb: Workbook):
    ws = wb.create_sheet("Budget")
    widths(ws, a=30)
    for i in range(MONTHS + 1):
        ws.column_dimensions[col(FIRST + i)].width = 12
    title(ws, "FY2026 operating budget")
    ws.freeze_panes = "B4"

    ws.cell(3, 1, "Line item").font = HDR_FONT
    ws.cell(3, 1).fill = HDR_FILL
    for i in range(MONTHS):
        c = ws.cell(3, FIRST + i, f"=EOMONTH(Assumptions!$B$16,{i})")
        c.number_format = MONTH
        c.font = HDR_FONT
        c.fill = HDR_FILL
    total_col = FIRST + MONTHS
    c = ws.cell(3, total_col, "FY total")
    c.font = HDR_FONT
    c.fill = HDR_FILL

    for row, label, _kind in LINES:
        ws.cell(row, 1, label)
    for row in (7, 11, 12, 22, 24):
        ws.cell(row, 1).font = BOLD
    ws.cell(13, 1, "Gross margin")
    ws.cell(25, 1, "EBITDA margin")
    ws.cell(27, 1, "YTD revenue").font = BOLD
    ws.cell(28, 1, "YTD EBITDA").font = BOLD
    ws.cell(14, 1, "Operating expenses").font = BOLD
    ws.cell(14, 1).fill = SUB_FILL

    for i in range(MONTHS):
        c = col(FIRST + i)
        prev = col(FIRST + i - 1)
        f = {
            5: f"=ROUND(Assumptions!$B$3*(1+Assumptions!$B$5)^{i}*Assumptions!$B${21 + i},1)",
            6: f"=ROUND(Assumptions!$B$4*(1+Assumptions!$B$6)^{i},1)",
            7: f"=SUM({c}5:{c}6)",
            9: f"=-ROUND({c}5*Assumptions!$B$7,1)",
            10: f"=-ROUND({c}6*Assumptions!$B$8,1)",
            11: f"=SUM({c}9:{c}10)",
            12: f"={c}7+{c}11",
            13: f'=IFERROR({c}12/{c}7,"")',
            15: f"=-ROUND(Assumptions!$C${21 + i}*Assumptions!$B$15/12/1000,1)",
            16: f"=ROUND({c}15*Assumptions!$B$14,1)",
            17: f"=-ROUND({c}7*Assumptions!$B$9,1)",
            18: "=-Assumptions!$B$11",
            19: "=-Assumptions!$B$12",
            20: f"=-ROUND({c}7*Assumptions!$B$10,1)",
            21: "=-Assumptions!$B$13",
            22: f"=SUM({c}15:{c}21)",
            24: f"={c}12+{c}22",
            25: f'=IFERROR({c}24/{c}7,"")',
            # The running total. Anchored left edge, sliding right edge — the
            # shape that turns one column into O(months^2) precedent cells.
            27: f"=SUM($B$7:{c}7)",
            28: f"=SUM($B$24:{c}24)",
        }
        for row, formula in f.items():
            cell = ws.cell(row, FIRST + i, formula)
            cell.number_format = PCT if row in (13, 25) else MONEY
            if row in (7, 11, 12, 22, 24, 27, 28):
                cell.font = BOLD
        # Month-over-month growth, blank in the first column rather than #DIV/0!.
        g = ws.cell(30, FIRST + i, "" if i == 0 else f'=IFERROR({c}7/{prev}7-1,"")')
        g.number_format = PCT
    ws.cell(30, 1, "Revenue growth, MoM")

    tc = col(total_col)
    for row, _label, _kind in LINES:
        cell = ws.cell(row, total_col, f"=SUM(B{row}:M{row})")
        cell.number_format = MONEY
        cell.font = BOLD
        cell.border = TOP_BORDER
    ws.cell(13, total_col, f"=IFERROR({tc}12/{tc}7,\"\")").number_format = PCT
    ws.cell(25, total_col, f"=IFERROR({tc}24/{tc}7,\"\")").number_format = PCT
    ws.cell(24, total_col).border = TOTAL_BORDER

    ws.cell(32, 1, "Sanity: FY total equals December YTD").font = BOLD
    ws.cell(32, 2, f'=IF(ROUND({tc}7-M27,2)=0,"tie","OUT BY "&TEXT({tc}7-M27,"#,##0.00"))')
    return ws


def actuals(wb: Workbook):
    """Literal actuals. Deterministic noise so the corpus is reproducible."""
    ws = wb.create_sheet("Actuals")
    widths(ws, a=30)
    for i in range(MONTHS):
        ws.column_dimensions[col(FIRST + i)].width = 12
    title(ws, "FY2026 actuals (posted)")
    ws.freeze_panes = "B4"

    ws.cell(3, 1, "Line item").font = HDR_FONT
    ws.cell(3, 1).fill = HDR_FILL
    for i in range(MONTHS):
        c = ws.cell(3, FIRST + i, f"=Budget!{col(FIRST + i)}3")
        c.number_format = MONTH
        c.font = HDR_FONT
        c.fill = HDR_FILL

    rng = random.Random(11)
    base = {
        5: 850.0, 6: 320.0, 9: -238.0, 10: -176.0,
        15: -757.0, 16: -167.0, 17: -164.0, 18: -68.0,
        19: -41.0, 20: -21.0, 21: -25.0,
    }
    growth = {5: 0.021, 6: 0.015, 9: 0.021, 10: 0.015, 15: 0.014,
              16: 0.014, 17: 0.021, 18: 0.0, 19: 0.0, 20: 0.021, 21: 0.0}
    for row, label, _kind in LINES:
        ws.cell(row, 1, label)
        if row not in base:
            continue
        for i in range(MONTHS):
            drift = 1 + growth[row] * i
            noise = 1 + rng.uniform(-0.09, 0.11)
            cell = ws.cell(row, FIRST + i, round(base[row] * drift * noise, 1))
            cell.number_format = MONEY

    # Subtotals stay formulas — actuals sheets are exports plus a few sums.
    for i in range(MONTHS):
        c = col(FIRST + i)
        for row, formula in {
            7: f"=SUM({c}5:{c}6)",
            11: f"=SUM({c}9:{c}10)",
            12: f"={c}7+{c}11",
            22: f"=SUM({c}15:{c}21)",
            24: f"={c}12+{c}22",
        }.items():
            cell = ws.cell(row, FIRST + i, formula)
            cell.number_format = MONEY
            cell.font = BOLD
    return ws


def variance(wb: Workbook):
    ws = wb.create_sheet("Variance")
    widths(ws, a=30, b=14, c=14, d=14, e=10, f=12, g=26)
    title(ws, "Full-year variance — actual vs budget")
    note(ws, 2, "Favourable means better than plan: more income, or less cost.")
    header_row(ws, 4, ["Line item", "Budget", "Actual", "Variance", "Var %", "F / U", "Flag"])

    for i, (row, label, kind) in enumerate(LINES):
        r = 5 + i
        ws.cell(r, 1, label)
        ws.cell(r, 2, f"=Budget!N{row}").number_format = MONEY
        ws.cell(r, 3, f"=SUM(Actuals!B{row}:M{row})").number_format = MONEY
        ws.cell(r, 4, f"=C{r}-B{r}").number_format = MONEY
        ws.cell(r, 5, f'=IFERROR(D{r}/ABS(B{r}),"n/a")').number_format = PCT
        # Cost lines are stored negative, so a variance above plan is favourable
        # for both kinds — but only because of the sign convention, which is
        # exactly the sort of thing worth stating in the formula.
        ws.cell(r, 6, f'=IF(D{r}=0,"—",IF(D{r}>0,"F","U"))')
        ws.cell(
            r, 7,
            f'=IF(AND(ABS(D{r})>Assumptions!$B$17,ABS(IFERROR(D{r}/B{r},0))>Assumptions!$B$18),'
            f'"REVIEW — "&TEXT(ABS(D{r}),"#,##0")&" "&IF(D{r}<0,"adverse","favourable"),"")'
        )
        if row in (7, 11, 12, 22, 24):
            ws.cell(r, 1).font = BOLD

    last = 4 + len(LINES)
    ws.cell(last + 2, 1, "Lines flagged for review").font = BOLD
    ws.cell(last + 2, 2, f'=COUNTIF(G5:G{last},"REVIEW*")').number_format = INT
    ws.cell(last + 3, 1, "Favourable lines").font = BOLD
    ws.cell(last + 3, 2, f'=COUNTIF(F5:F{last},"F")').number_format = INT
    ws.cell(last + 4, 1, "Total adverse variance").font = BOLD
    ws.cell(last + 4, 2, f'=SUMIF(D5:D{last},"<0")').number_format = MONEY
    ws.cell(last + 5, 1, "Worst line").font = BOLD
    ws.cell(last + 5, 2, f"=INDEX(A5:A{last},MATCH(MIN(D5:D{last}),D5:D{last},0))")
    ws.cell(last + 6, 1, "EBITDA variance vs plan").font = BOLD
    ws.cell(last + 6, 2, f'=TEXT(D{4 + len(LINES)},"#,##0.0")&" ("&TEXT(E{4 + len(LINES)},"0.0%")&")"')
    return ws


def build() -> Workbook:
    wb = Workbook()
    assumptions(wb)
    budget(wb)
    actuals(wb)
    variance(wb)
    return wb
