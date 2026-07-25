"""M08 — 360-month mortgage amortisation with overpayments.

A full monthly schedule for a thirty-year loan, with a fixed extra principal
payment that pays it off early, and a summary that reconciles the actual
schedule against the closed-form functions.

What it stresses:

  chain depth     each month's opening balance is the previous month's closing
                  balance, 360 links deep in a single column. Evaluation order
                  here is not a graph-shaped problem, it is a rope, and a
                  recursive evaluator overflows its stack on exactly this shape
  financial fns   PMT, IPMT, PPMT, CUMIPMT, CUMPRINC, RATE, NPER, EFFECT and
                  NOMINAL, several of which are iterative solvers rather than
                  closed forms
  rounding        every line is rounded to cents, so the closing balance is the
                  accumulation of 360 rounding decisions. A ROUND that differs
                  in the last bit compounds visibly by month 300
  early payoff    once the balance reaches zero the remaining rows must stay at
                  zero rather than going negative or dividing by it
"""

from __future__ import annotations

from openpyxl import Workbook

from common import (
    BOLD, DATE, HDR_FILL, HDR_FONT, INPUT_FONT, INT, MONEY2, NUM, PCT, PCT2,
    SUB_FILL, col, header_row, note, title, widths,
)

TERMS = 360
TOP = 5  # first schedule row


def loan(wb: Workbook):
    ws = wb.active
    ws.title = "Loan"
    widths(ws, a=36, b=16)
    title(ws, "Loan terms")

    rows = [
        ("Principal", 850000.0, MONEY2),
        ("Annual interest rate", 0.0575, PCT2),
        ("Term, years", 30, INT),
        ("Payments per year", 12, INT),
        ("Extra principal, per month", 250.0, MONEY2),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value)
        c.number_format = fmt
        c.font = INPUT_FONT
    ws.cell(8, 1, "First payment date").font = BOLD
    ws.cell(8, 2, "=DATE(2026,3,1)").number_format = DATE

    derived = [
        ("Periodic rate", "=B4/B6", "0.0000000"),
        ("Number of periods", "=B5*B6", INT),
        ("Scheduled payment", "=-PMT(B10,B11,B3)", MONEY2),
        ("Total scheduled interest", "=-CUMIPMT(B10,B11,B3,1,B11,0)", MONEY2),
        ("Total scheduled principal", "=-CUMPRINC(B10,B11,B3,1,B11,0)", MONEY2),
        ("Total scheduled payments", "=B12*B11", MONEY2),
        ("Interest plus principal ties",
         '=IF(ABS(B13+B14-B15)<1,"ties","OUT BY "&TEXT(B13+B14-B15,"#,##0.00"))', "General"),
        ("Interest in year 1", "=-CUMIPMT(B10,B11,B3,1,12,0)", MONEY2),
        ("Interest in year 30", "=-CUMIPMT(B10,B11,B3,349,360,0)", MONEY2),
        ("Effective annual rate", "=EFFECT(B4,B6)", PCT2),
        ("Nominal from effective (round trip)", "=NOMINAL(B19,B6)", PCT2),
        ("Round trip returns the input",
         '=IF(ABS(B20-B4)<0.0000001,"yes","NO")', "General"),
        ("Periods to payoff with overpayment",
         "=NPER(B10,-(B12+B7),B3)", NUM),
        ("Rate implied by the scheduled payment",
         "=RATE(B11,-B12,B3)*B6", PCT2),
        ("Implied rate returns the input",
         '=IF(ABS(B23-B4)<0.000001,"yes","NO")', "General"),
    ]
    for i, (label, formula, fmt) in enumerate(derived, start=10):
        ws.cell(i, 1, label).font = BOLD
        ws.cell(i, 2, formula).number_format = fmt
    return ws


def schedule(wb: Workbook):
    ws = wb.create_sheet("Schedule")
    title(ws, "Amortisation schedule")
    note(ws, 2, f"{TERMS} scheduled periods. The loan pays off early because of the overpayment.")
    ws.freeze_panes = "B5"

    cols = [
        ("Period", 8), ("Payment date", 13), ("Opening balance", 15),
        ("Payment", 13), ("Interest", 13), ("Principal", 13),
        ("Extra principal", 14), ("Closing balance", 15),
        ("Cumulative interest", 17), ("Cumulative principal", 18),
        ("Paid off %", 11), ("Year", 8),
        ("Scheduled interest (IPMT)", 20), ("Scheduled principal (PPMT)", 21),
        ("Interest saved to date", 18),
    ]
    header_row(ws, 4, [c[0] for c in cols])
    for i, (_label, w) in enumerate(cols):
        ws.column_dimensions[col(1 + i)].width = w

    for k in range(TERMS):
        r = TOP + k
        p = r - 1  # previous row
        ws.cell(r, 1, k + 1).number_format = INT
        ws.cell(r, 2, f"=EDATE(Loan!$B$8,{k})").number_format = DATE
        ws.cell(r, 3, ("=Loan!$B$3" if k == 0 else f"=H{p}")).number_format = MONEY2
        # Once the balance is gone every downstream line has to stay at zero
        # rather than turning negative or dividing by it.
        ws.cell(r, 4,
                f"=IF(C{r}<=0,0,MIN(Loan!$B$12,ROUND(C{r}*(1+Loan!$B$10),2)))"
                ).number_format = MONEY2
        ws.cell(r, 5, f"=ROUND(C{r}*Loan!$B$10,2)").number_format = MONEY2
        ws.cell(r, 6, f"=ROUND(D{r}-E{r},2)").number_format = MONEY2
        ws.cell(r, 7,
                f"=IF(C{r}-F{r}<=0,0,MIN(Loan!$B$7,ROUND(C{r}-F{r},2)))").number_format = MONEY2
        ws.cell(r, 8, f"=MAX(0,ROUND(C{r}-F{r}-G{r},2))").number_format = MONEY2
        ws.cell(r, 9, (f"=E{r}" if k == 0 else f"=I{p}+E{r}")).number_format = MONEY2
        ws.cell(r, 10, (f"=F{r}+G{r}" if k == 0 else f"=J{p}+F{r}+G{r}")).number_format = MONEY2
        ws.cell(r, 11, f"=1-H{r}/Loan!$B$3").number_format = PCT
        ws.cell(r, 12, f"=YEAR(B{r})").number_format = "0"
        ws.cell(r, 13,
                f"=-IPMT(Loan!$B$10,A{r},Loan!$B$11,Loan!$B$3)").number_format = MONEY2
        ws.cell(r, 14,
                f"=-PPMT(Loan!$B$10,A{r},Loan!$B$11,Loan!$B$3)").number_format = MONEY2
        ws.cell(r, 15,
                (f"=M{r}-E{r}" if k == 0 else f"=O{p}+M{r}-E{r}")).number_format = MONEY2
    return ws


def summary(wb: Workbook):
    ws = wb.create_sheet("Summary")
    widths(ws, a=38, b=16, c=16, d=16, e=16)
    title(ws, "Overpayment outcome")

    last = TOP + TERMS - 1
    closing = f"Schedule!$H${TOP}:$H${last}"
    interest = f"Schedule!$E${TOP}:$E${last}"
    extra = f"Schedule!$G${TOP}:$G${last}"
    payment = f"Schedule!$D${TOP}:$D${last}"
    period = f"Schedule!$A${TOP}:$A${last}"
    year = f"Schedule!$L${TOP}:$L${last}"

    rows = [
        ("Payoff period", f"=MATCH(0,{closing},0)", INT),
        ("Payoff date",
         f"=INDEX(Schedule!$B${TOP}:$B${last},MATCH(0,{closing},0))", DATE),
        ("Periods saved", "=Loan!$B$11-B4", INT),
        ("Years saved", "=B6/12", NUM),
        ("Interest actually paid", f"=SUM({interest})", MONEY2),
        ("Interest if never overpaid", "=Loan!$B$13", MONEY2),
        ("Interest saved", "=B9-B8", MONEY2),
        ("Interest saved, %", "=B10/B9", PCT),
        ("Extra principal paid", f"=SUM({extra})", MONEY2),
        ("Return on the overpayment", "=B10/B12", "0.00"),
        ("Total cash paid", f"=SUM({payment})+SUM({extra})", MONEY2),
        ("Principal repaid", f"=B14-SUM({interest})", MONEY2),
        ("Principal repaid ties to the loan",
         '=IF(ABS(B15-Loan!$B$3)<1,"ties","OUT BY "&TEXT(B15-Loan!$B$3,"#,##0.00"))',
         "General"),
        ("NPER agrees with the schedule",
         '=IF(ABS(Loan!$B$22-B4)<1,"agrees","differs by "&TEXT(Loan!$B$22-B4,"0.0")&" periods")',
         "General"),
        ("Largest single interest charge", f"=MAX({interest})", MONEY2),
        ("Average payment", f'=AVERAGEIF({payment},">0")', MONEY2),
        ("Periods where the payment was capped",
         f'=COUNTIFS({payment},">0",{payment},"<"&Loan!$B$12)', INT),
    ]
    for i, (label, formula, fmt) in enumerate(rows, start=4):
        ws.cell(i, 1, label).font = BOLD
        ws.cell(i, 2, formula).number_format = fmt

    base = 23
    ws.cell(base, 1, "Interest and principal by calendar year").font = BOLD
    ws.cell(base, 1).fill = SUB_FILL
    header_row(ws, base + 1,
               ["Year", "Interest", "Principal", "Extra", "Closing balance", "% repaid"])
    for i in range(31):
        r = base + 2 + i
        ws.cell(r, 1, 2026 + i).number_format = "0"
        ws.cell(r, 2, f"=SUMIF({year},$A{r},{interest})").number_format = MONEY2
        ws.cell(r, 3,
                f"=SUMIF({year},$A{r},Schedule!$F${TOP}:$F${last})").number_format = MONEY2
        ws.cell(r, 4, f"=SUMIF({year},$A{r},{extra})").number_format = MONEY2
        ws.cell(r, 5,
                f'=IFERROR(INDEX({closing},MATCH($A{r},{year},0)+COUNTIF({year},$A{r})-1),0)'
                ).number_format = MONEY2
        ws.cell(r, 6, f"=1-E{r}/Loan!$B$3").number_format = PCT
    yr_last = base + 32
    ws.cell(yr_last + 1, 1, "Total").font = BOLD
    for c in "BCD":
        ws.cell(yr_last + 1, ord(c) - 64,
                f"=SUM({c}{base + 2}:{c}{yr_last})").number_format = MONEY2
        ws.cell(yr_last + 1, ord(c) - 64).font = BOLD
    ws.cell(yr_last + 2, 1, "Yearly interest ties to the schedule").font = BOLD
    ws.cell(yr_last + 2, 2,
            f'=IF(ABS(B{yr_last + 1}-SUM({interest}))<0.05,"ties","OUT BY "'
            f'&TEXT(B{yr_last + 1}-SUM({interest}),"#,##0.00"))')
    ws.cell(yr_last + 3, 1, "First year with no interest").font = BOLD
    ws.cell(yr_last + 3, 2,
            f'=IFERROR(INDEX($A${base + 2}:$A${yr_last},MATCH(0,$B${base + 2}:$B${yr_last},0)),'
            f'"loan never pays off")').number_format = "0"
    return ws


def build() -> Workbook:
    wb = Workbook()
    loan(wb)
    schedule(wb)
    summary(wb)
    return wb
