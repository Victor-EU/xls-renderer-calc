"""M02 — Discounted cash flow valuation with a sensitivity grid.

A five-year unlevered FCF forecast, a WACC built from CAPM, terminal value by
both Gordon growth and exit multiple, and a WACC × terminal-growth sensitivity
table underneath.

What it stresses: NPV/XNPV/XIRR against dated flows, `^` precedence and
associativity in discount factors, and above all the sensitivity grid — 63 cells
that each recompute the whole valuation through mixed anchoring (`$A12` down,
`C$10` across). Mixed anchoring is where a shared-formula reader that hands back
the master text unchanged produces 63 copies of one number and looks fine.
"""

from __future__ import annotations

from openpyxl import Workbook

from common import (
    BOLD, DATE, Grid, HDR_FILL, HDR_FONT, INPUT_FONT, MONEY, MONEY2, MULT,
    NUM, PCT, PCT2, SUB_FILL, TOTAL_BORDER, col, header_row, note, title, widths,
)

YEARS = 5
FIRST = 2


def inputs(wb: Workbook):
    ws = wb.active
    ws.title = "Inputs"
    widths(ws, a=34, b=14)
    title(ws, "Valuation inputs")
    note(ws, 2, "Ticker NVEX · valuation date 30 Jun 2026 · USD millions except per share")

    rows = [
        ("LTM revenue", 4820.0, MONEY),
        ("LTM EBITDA margin", 0.223, PCT),
        ("Revenue growth, year 1", 0.140, PCT),
        ("Revenue growth, year 5", 0.060, PCT),
        ("Terminal EBITDA margin", 0.265, PCT),
        ("D&A, % of revenue", 0.041, PCT),
        ("Capex, % of revenue", 0.052, PCT),
        ("Net working capital, % of revenue", 0.086, PCT),
        ("Cash tax rate", 0.240, PCT),
        ("Risk-free rate", 0.042, PCT2),
        ("Equity risk premium", 0.055, PCT2),
        ("Levered beta", 1.18, "0.00"),
        ("Pre-tax cost of debt", 0.061, PCT2),
        ("Debt / total capital", 0.280, PCT),
        ("Terminal growth rate", 0.025, PCT2),
        ("Exit EBITDA multiple", 11.5, MULT),
        ("Net debt", 1240.0, MONEY),
        ("Diluted shares outstanding", 182.4, NUM),
        ("Current share price", 34.20, MONEY2),
    ]
    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(i, 1, label)
        c = ws.cell(i, 2, value)
        c.number_format = fmt
        c.font = INPUT_FONT

    ws.cell(22, 1, "Valuation date").font = BOLD
    ws.cell(22, 2, "=DATE(2026,6,30)").number_format = DATE

    ws.cell(24, 1, "Cost of equity (CAPM)").font = BOLD
    ws.cell(24, 2, "=B12+B14*B13").number_format = PCT2
    ws.cell(25, 1, "After-tax cost of debt").font = BOLD
    ws.cell(25, 2, "=B15*(1-B11)").number_format = PCT2
    ws.cell(26, 1, "WACC").font = BOLD
    ws.cell(26, 2, "=B16*B25+(1-B16)*B24").number_format = PCT2
    ws.cell(26, 2).border = TOTAL_BORDER
    note(ws, 27, "WACC feeds the DCF and anchors the sensitivity grid.")
    return ws


def forecast(wb: Workbook):
    ws = wb.create_sheet("Forecast")
    widths(ws, a=32)
    for i in range(YEARS):
        ws.column_dimensions[col(FIRST + i)].width = 13
    title(ws, "Unlevered free cash flow forecast")
    ws.freeze_panes = "B5"

    ws.cell(4, 1, "Year ending").font = HDR_FONT
    ws.cell(4, 1).fill = HDR_FILL
    for i in range(YEARS):
        c = ws.cell(4, FIRST + i, f"=EDATE(Inputs!$B$22,{12 * (i + 1)})")
        c.number_format = DATE
        c.font = HDR_FONT
        c.fill = HDR_FILL

    g = Grid(ws, FIRST, YEARS, start_row=6)
    g.line("growth", "Revenue growth",
           lambda i, c: f"=Inputs!$B$5+(Inputs!$B$6-Inputs!$B$5)*{i}/{YEARS - 1}", PCT)
    g.line("revenue", "Revenue",
           lambda i, c: (f"=Inputs!$B$3*(1+{c}{g.R['growth']})" if i == 0
                         else f"={col(FIRST + i - 1)}{g.R['revenue']}*(1+{c}{g.R['growth']})"),
           bold=True)
    g.line("margin", "EBITDA margin",
           lambda i, c: f"=Inputs!$B$4+(Inputs!$B$7-Inputs!$B$4)*{i + 1}/{YEARS}", PCT)
    g.line("ebitda", "EBITDA",
           lambda i, c: f"=ROUND({c}{g.R['revenue']}*{c}{g.R['margin']},1)", bold=True)
    g.line("da", "Depreciation & amortisation",
           lambda i, c: f"=-ROUND({c}{g.R['revenue']}*Inputs!$B$8,1)")
    g.line("ebit", "EBIT",
           lambda i, c: f"={c}{g.R['ebitda']}+{c}{g.R['da']}", bold=True)
    g.line("tax", "Cash taxes",
           lambda i, c: f"=-MAX(0,ROUND({c}{g.R['ebit']}*Inputs!$B$11,1))")
    g.line("nopat", "NOPAT", lambda i, c: f"={c}{g.R['ebit']}+{c}{g.R['tax']}", bold=True)
    g.line("addback", "Add back D&A", lambda i, c: f"=-{c}{g.R['da']}")
    g.line("capex", "Capital expenditure",
           lambda i, c: f"=-ROUND({c}{g.R['revenue']}*Inputs!$B$9,1)")
    g.line("nwc", "Net working capital balance",
           lambda i, c: f"=ROUND({c}{g.R['revenue']}*Inputs!$B$10,1)")
    g.line("dnwc", "(Increase) in NWC",
           lambda i, c: (f"=-({c}{g.R['nwc']}-ROUND(Inputs!$B$3*Inputs!$B$10,1))" if i == 0
                         else f"=-({c}{g.R['nwc']}-{col(FIRST + i - 1)}{g.R['nwc']})"))
    g.gap()
    g.line("fcf", "Unlevered free cash flow",
           lambda i, c: (f"={c}{g.R['nopat']}+{c}{g.R['addback']}"
                         f"+{c}{g.R['capex']}+{c}{g.R['dnwc']}"), bold=True)
    g.line("fcfmargin", "FCF margin",
           lambda i, c: f'=IFERROR({c}{g.R["fcf"]}/{c}{g.R["revenue"]},"")', PCT)
    g.line("conv", "Cash conversion (FCF / EBITDA)",
           lambda i, c: f'=IFERROR({c}{g.R["fcf"]}/{c}{g.R["ebitda"]},"n/m")', PCT)

    g.render()
    for r in (g.R["fcf"],):
        for i in range(YEARS):
            ws.cell(r, FIRST + i).border = TOTAL_BORDER

    ws.cell(g.r + 1, 1, "Revenue CAGR, LTM to year 5").font = BOLD
    ws.cell(g.r + 1, 2,
            f"=({g.ref('revenue', YEARS - 1)}/Inputs!$B$3)^(1/{YEARS})-1").number_format = PCT
    ws.cell(g.r + 2, 1, "Sum of forecast FCF").font = BOLD
    ws.cell(g.r + 2, 2, f"=SUM({g.span('fcf')})").number_format = MONEY
    return ws, g


def dcf(wb: Workbook, fg: Grid):
    ws = wb.create_sheet("DCF")
    widths(ws, a=34, b=14, c=13, d=13, e=13, f=13, g=13)
    title(ws, "Discounted cash flow")
    ws.cell(4, 1, "Year").font = HDR_FONT
    ws.cell(4, 1).fill = HDR_FILL
    for i in range(YEARS):
        c = ws.cell(4, FIRST + i, f"=YEAR(Forecast!{col(FIRST + i)}4)")
        c.number_format = "0"
        c.font = HDR_FONT
        c.fill = HDR_FILL

    g = Grid(ws, FIRST, YEARS, start_row=6)
    g.line("period", "Discount period", lambda i, c: f"={i + 1}", "0")
    g.line("factor", "Discount factor",
           lambda i, c: f"=1/(1+Inputs!$B$26)^{c}{g.R['period']}", "0.0000")
    g.line("fcf", "Unlevered FCF", lambda i, c: f"=Forecast!{c}{fg.R['fcf']}")
    g.line("pv", "PV of FCF", lambda i, c: f"={c}{g.R['fcf']}*{c}{g.R['factor']}", bold=True)
    g.gap()

    g.render()

    last = col(FIRST + YEARS - 1)
    g.rule("sumpv", "Sum of PV, explicit period", f"=SUM({g.span('pv')})")
    g.rule("crosscheck", "Cross-check with NPV()",
           f"=NPV(Inputs!$B$26,{g.span('fcf')})")
    g.rule("npvdelta", 'NPV cross-check ties',
           f'=IF(ABS({g.ref("sumpv", 0)}-{g.ref("crosscheck", 0)})<0.05,"ties",'
           f'"differs by "&TEXT({g.ref("sumpv", 0)}-{g.ref("crosscheck", 0)},"0.000"))', "General")
    g.gap()
    g.rule("tv_gordon", "Terminal value — Gordon growth",
           f"={last}{g.R['fcf']}*(1+Inputs!$B$17)/(Inputs!$B$26-Inputs!$B$17)")
    g.rule("pv_gordon", "PV of terminal value",
           f"={g.ref('tv_gordon', 0)}*{last}{g.R['factor']}")
    g.rule("tv_exit", "Terminal value — exit multiple",
           f"=Forecast!{last}{fg.R['ebitda']}*Inputs!$B$18")
    g.rule("pv_exit", "PV of terminal value",
           f"={g.ref('tv_exit', 0)}*{last}{g.R['factor']}")
    g.gap()
    g.rule("ev_gordon", "Enterprise value — Gordon",
           f"={g.ref('sumpv', 0)}+{g.ref('pv_gordon', 0)}")
    g.rule("ev_exit", "Enterprise value — exit multiple",
           f"={g.ref('sumpv', 0)}+{g.ref('pv_exit', 0)}")
    g.rule("ev", "Enterprise value — average of methods",
           f"=AVERAGE({g.ref('ev_gordon', 0)},{g.ref('ev_exit', 0)})")
    g.rule("netdebt", "Less: net debt", "=-Inputs!$B$19")
    g.rule("equity", "Equity value", f"={g.ref('ev', 0)}+{g.ref('netdebt', 0)}")
    g.rule("pershare", "Value per share",
           f"={g.ref('equity', 0)}/Inputs!$B$20", MONEY2)
    g.rule("upside", "Upside to current price",
           f"={g.ref('pershare', 0)}/Inputs!$B$21-1", PCT)
    g.gap()
    g.rule("tvshare", "Terminal value as % of EV",
           f"={g.ref('pv_gordon', 0)}/{g.ref('ev_gordon', 0)}", PCT)
    g.rule("impmult", "Implied EV / LTM EBITDA",
           f"={g.ref('ev', 0)}/(Inputs!$B$3*Inputs!$B$4)", MULT)
    g.rule("call", "Recommendation",
           f'=IF({g.ref("upside", 0)}>0.20,"BUY",'
           f'IF({g.ref("upside", 0)}>0.05,"ACCUMULATE",'
           f'IF({g.ref("upside", 0)}>-0.10,"HOLD","REDUCE")))', "General")
    ws.cell(g.R["pershare"], FIRST).border = TOTAL_BORDER

    # Dated cash flows — the same valuation done with XNPV/XIRR, which are what
    # an analyst reaches for when the periods are not exactly annual.
    r = g.r + 2
    ws.cell(r, 1, "Dated flows (XNPV / XIRR cross-check)").font = BOLD
    ws.cell(r, 1).fill = SUB_FILL
    ws.cell(r + 1, 1, "Date")
    ws.cell(r + 1, FIRST, "=Inputs!$B$22").number_format = DATE
    for i in range(YEARS):
        ws.cell(r + 1, FIRST + 1 + i, f"=Forecast!{col(FIRST + i)}4").number_format = DATE
    ws.cell(r + 2, 1, "Cash flow")
    ws.cell(r + 2, FIRST, f"=-{g.ref('ev', 0)}").number_format = MONEY
    for i in range(YEARS):
        f = f"=Forecast!{col(FIRST + i)}{fg.R['fcf']}"
        if i == YEARS - 1:
            f += f"+{g.ref('tv_gordon', 0)}"
        ws.cell(r + 2, FIRST + 1 + i, f).number_format = MONEY
    flows = f"B{r + 2}:{col(FIRST + YEARS)}{r + 2}"
    dates = f"B{r + 1}:{col(FIRST + YEARS)}{r + 1}"
    ws.cell(r + 4, 1, "XNPV at WACC (should be ~0)").font = BOLD
    ws.cell(r + 4, 2, f"=XNPV(Inputs!$B$26,{flows},{dates})").number_format = "#,##0.00"
    ws.cell(r + 5, 1, "XIRR (should be ~WACC)").font = BOLD
    ws.cell(r + 5, 2, f"=XIRR({flows},{dates})").number_format = PCT2
    ws.cell(r + 6, 1, "IRR on annual flows").font = BOLD
    ws.cell(r + 6, 2, f"=IRR({flows})").number_format = PCT2
    ws.cell(r + 7, 1, "XIRR ties to WACC").font = BOLD
    ws.cell(r + 7, 2, f'=IF(ABS(B{r + 5}-Inputs!$B$26)<0.0005,"ties","differs")')
    return ws, g


def sensitivity(wb: Workbook, fg: Grid, dg: Grid):
    ws = wb.create_sheet("Sensitivity")
    widths(ws, a=16)
    title(ws, "Value per share — WACC vs terminal growth")
    note(ws, 2, "Each cell re-runs the whole valuation; the centre cell must tie to the DCF sheet.")

    waccs = 7
    growths = 5
    top = 5   # header row
    left = 1  # WACC column A

    ws.cell(top, left, "WACC \\ g").font = HDR_FONT
    ws.cell(top, left).fill = HDR_FILL
    for j in range(growths):
        c = ws.cell(top, 2 + j, f"=Inputs!$B$17+{(j - (growths - 1) // 2) * 0.005}")
        c.number_format = PCT2
        c.font = HDR_FONT
        c.fill = HDR_FILL
        ws.column_dimensions[col(2 + j)].width = 12
    for i in range(waccs):
        c = ws.cell(top + 1 + i, left, f"=Inputs!$B$26+{(i - (waccs - 1) // 2) * 0.005}")
        c.number_format = PCT2
        c.font = HDR_FONT
        c.fill = HDR_FILL

    # The full valuation written out per cell. Analysts really do write this;
    # it is also the cleanest way to exercise mixed anchoring at scale.
    for i in range(waccs):
        r = top + 1 + i
        for j in range(growths):
            terms = "+".join(
                f"Forecast!${col(FIRST + k)}${fg.R['fcf']}/(1+$A{r})^{k + 1}" for k in range(YEARS)
            )
            tv = (f"Forecast!${col(FIRST + YEARS - 1)}${fg.R['fcf']}*(1+{col(2 + j)}${top})"
                  f"/($A{r}-{col(2 + j)}${top})/(1+$A{r})^{YEARS}")
            cell = ws.cell(r, 2 + j,
                           f"=IF($A{r}<={col(2 + j)}${top},\"n/m\","
                           f"(({terms}+{tv})-Inputs!$B$19)/Inputs!$B$20)")
            cell.number_format = "#,##0.00"

    mid_r = top + 1 + (waccs - 1) // 2
    mid_c = 2 + (growths - 1) // 2
    base = top + waccs + 3
    ws.cell(base, 1, "Centre cell").font = BOLD
    ws.cell(base, 2, f"={col(mid_c)}{mid_r}").number_format = "#,##0.00"
    ws.cell(base + 1, 1, "DCF sheet (Gordon only)").font = BOLD
    ws.cell(base + 1, 2,
            f"=(DCF!B{dg.R['sumpv']}+DCF!B{dg.R['pv_gordon']}-Inputs!$B$19)"
            f"/Inputs!$B$20").number_format = "#,##0.00"
    ws.cell(base + 2, 1, "Grid ties to model").font = BOLD
    ws.cell(base + 2, 2,
            f'=IF(ABS(B{base}-B{base + 1})<0.01,"ties",'
            f'"OUT BY "&TEXT(B{base}-B{base + 1},"0.000"))')
    ws.cell(base + 4, 1, "Range across the grid").font = BOLD
    grid = f"B{top + 1}:{col(1 + growths)}{top + waccs}"
    ws.cell(base + 4, 2, f"=MAX({grid})-MIN({grid})").number_format = "#,##0.00"
    ws.cell(base + 5, 1, "Cells above current price").font = BOLD
    ws.cell(base + 5, 2, f"=COUNTIF({grid},\">\"&Inputs!$B$21)").number_format = "0"
    ws.cell(base + 6, 1, "Median outcome").font = BOLD
    ws.cell(base + 6, 2, f"=MEDIAN({grid})").number_format = "#,##0.00"
    return ws


def comps(wb: Workbook, dg: Grid):
    ws = wb.create_sheet("Comps")
    widths(ws, a=26, b=14, c=14, d=14, e=14, f=12)
    title(ws, "Trading comparables")
    header_row(ws, 4, ["Company", "EV ($m)", "EBITDA ($m)", "EV/EBITDA", "Growth", "Rank"])
    peers = [
        ("Arden Systems", 14820.0, 1410.0, 0.118),
        ("Boreal Data", 8940.0, 705.0, 0.164),
        ("Corvine Labs", 22310.0, 2480.0, 0.091),
        ("Delmar Networks", 5120.0, 612.0, 0.073),
        ("Estrel Software", 31450.0, 2210.0, 0.209),
        ("Fenwick Cloud", 11060.0, 998.0, 0.142),
    ]
    for i, (name, ev, ebitda, growth) in enumerate(peers):
        r = 5 + i
        ws.cell(r, 1, name)
        ws.cell(r, 2, ev).number_format = MONEY
        ws.cell(r, 3, ebitda).number_format = MONEY
        ws.cell(r, 4, f"=B{r}/C{r}").number_format = MULT
        ws.cell(r, 5, growth).number_format = PCT
        ws.cell(r, 6, f"=RANK.EQ(D{r},$D$5:$D$10,0)").number_format = "0"
    last = 4 + len(peers)
    stats = [
        ("Median", f"=MEDIAN(D5:D{last})"),
        ("Mean", f"=AVERAGE(D5:D{last})"),
        ("25th percentile", f"=PERCENTILE.INC(D5:D{last},0.25)"),
        ("75th percentile", f"=PERCENTILE.INC(D5:D{last},0.75)"),
        ("Std deviation", f"=STDEV.S(D5:D{last})"),
        ("Growth-adjusted median", f"=MEDIAN(D5:D{last})/MEDIAN(E5:E{last})"),
    ]
    for i, (label, formula) in enumerate(stats):
        r = last + 2 + i
        ws.cell(r, 1, label).font = BOLD
        ws.cell(r, 4, formula).number_format = MULT
    base = last + 2 + len(stats) + 1
    ws.cell(base, 1, "Implied EV at median multiple").font = BOLD
    ws.cell(base, 4, f"=D{last + 2}*Inputs!$B$3*Inputs!$B$4").number_format = MONEY
    ws.cell(base + 1, 1, "Implied value per share").font = BOLD
    ws.cell(base + 1, 4, f"=(D{base}-Inputs!$B$19)/Inputs!$B$20").number_format = MONEY2
    ws.cell(base + 2, 1, "DCF vs comps").font = BOLD
    ws.cell(base + 2, 4,
            f'=IF(D{base + 1}>DCF!B{dg.R["pershare"]},"comps richer","DCF richer")')
    ws.cell(base + 3, 1, "Peers above our implied multiple").font = BOLD
    ws.cell(base + 3, 4, f'=COUNTIF(D5:D{last},">"&D{last + 2})').number_format = "0"
    ws.cell(base + 4, 1, "Weighted multiple (by EBITDA)").font = BOLD
    ws.cell(base + 4, 4,
            f"=SUMPRODUCT(D5:D{last},C5:C{last})/SUM(C5:C{last})").number_format = MULT
    return ws


def build() -> Workbook:
    wb = Workbook()
    inputs(wb)
    _, fg = forecast(wb)
    _, dg = dcf(wb, fg)
    sensitivity(wb, fg, dg)
    comps(wb, dg)
    return wb
