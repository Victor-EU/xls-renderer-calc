"""Shared plumbing for the eval corpus.

Two jobs, and deliberately nothing else — the model scripts themselves must read
like something an agent wrote in one pass, because that is what they are
standing in for.

    styling      the fonts, fills and number formats an agent applies without
                 thinking about it. Kept here so the model files stay about the
                 model.

    to_ooxml()   openpyxl writes formula text verbatim, but Excel stores every
                 function added after 2007 under the `_xlfn.` namespace. Without
                 the prefix LibreOffice reads `XLOOKUP` as an unknown name and
                 returns #NAME?, which in the report looks exactly like an engine
                 gap and is really a fixture bug. Applied centrally at save time
                 so no model script has to know about it.
"""

from __future__ import annotations

import re

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- styling ---

NAVY = "1F3864"
SLATE = "44546A"
PALE = "EEF2F8"

HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=PALE)
TITLE = Font(bold=True, size=13, color=NAVY)
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color=SLATE, size=9)
INPUT_FONT = Font(color="0000CC")  # the blue-is-an-input convention

TOP_BORDER = Border(top=Side(style="thin"))
TOTAL_BORDER = Border(top=Side(style="thin"), bottom=Side(style="double"))

MONEY = "#,##0.0;[Red](#,##0.0)"
MONEY0 = "#,##0;[Red](#,##0)"
MONEY2 = "$#,##0.00;[Red]($#,##0.00)"
PCT = "0.0%"
PCT2 = "0.00%"
MULT = '0.00"x"'
NUM = "#,##0.0"
INT = "#,##0"
DATE = "yyyy-mm-dd"
MONTH = "mmm-yy"


def title(ws, text: str, cell: str = "A1") -> None:
    ws[cell] = text
    ws[cell].font = TITLE


def note(ws, row: int, text: str, col: int = 1) -> None:
    c = ws.cell(row, col, text)
    c.font = ITALIC


def header_row(ws, row: int, labels, start_col: int = 1, width: int | None = None) -> None:
    for i, label in enumerate(labels):
        c = ws.cell(row, start_col + i, label)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if width:
            ws.column_dimensions[get_column_letter(start_col + i)].width = width


def widths(ws, **cols) -> None:
    for letter, w in cols.items():
        ws.column_dimensions[letter.upper()].width = w


def col(i: int) -> str:
    """1-based column index to letter — models index columns, not letters."""
    return get_column_letter(i)


class Grid:
    """A schedule laid out as labelled rows across N period columns.

    Rows are addressed by name rather than by number, because a model that
    renumbers itself every time a line is inserted is how row-reference bugs get
    written — including into eval fixtures, where they would masquerade as engine
    failures.
    """

    def __init__(self, ws, first_col: int, periods: int, start_row: int = 5):
        self.ws = ws
        self.first = first_col
        self.n = periods
        self.r = start_row
        self.R: dict[str, int] = {}
        self._pending: list[tuple] = []

    def ref(self, key: str, i: int, absolute: bool = False) -> str:
        """`C12`-style reference to another line in this grid, period `i`."""
        c = get_column_letter(self.first + i)
        return f"${c}${self.R[key]}" if absolute else f"{c}{self.R[key]}"

    def row_of(self, key: str) -> int:
        return self.R[key]

    def span(self, key: str) -> str:
        """`B12:F12` — the whole line, for a SUM or an IRR."""
        a = get_column_letter(self.first)
        b = get_column_letter(self.first + self.n - 1)
        return f"{a}{self.R[key]}:{b}{self.R[key]}"

    def line(self, key, label, fn=None, fmt=MONEY, bold=False, indent=0, fill=None):
        r = self.r
        self.r += 1
        self.R[key] = r
        c = self.ws.cell(r, 1, ("    " * indent) + label)
        if bold:
            c.font = BOLD
        if fill:
            c.fill = fill
        if fn is not None:
            # Deferred: a line may reference one defined below it — a closing
            # balance feeding next period's opening balance is the normal shape
            # of a schedule, and eager evaluation cannot express it.
            self._pending.append((r, fn, fmt, bold))
        return r

    def render(self) -> "Grid":
        """Write every deferred line. Call once the whole grid is laid out."""
        for r, fn, fmt, bold in self._pending:
            for i in range(self.n):
                value = fn(i, get_column_letter(self.first + i))
                if value is None:
                    continue
                cell = self.ws.cell(r, self.first + i, value)
                cell.number_format = fmt
                if bold:
                    cell.font = BOLD
        self._pending.clear()
        return self

    def gap(self, k: int = 1):
        self.r += k

    def rule(self, key: str, label: str, formula: str, fmt=MONEY):
        """A single-cell line — a total or a check, not a per-period row."""
        r = self.r
        self.r += 1
        self.R[key] = r
        self.ws.cell(r, 1, label).font = BOLD
        cell = self.ws.cell(r, self.first, formula)
        cell.number_format = fmt
        cell.font = BOLD
        return r


# ----------------------------------------------------------- ooxml naming ---

XLFN = {
    "IFS", "SWITCH", "MAXIFS", "MINIFS", "TEXTJOIN", "CONCAT", "XOR", "IFNA",
    "XLOOKUP", "XMATCH", "DAYS", "RRI", "PDURATION", "UNICHAR", "UNICODE",
    "NUMBERVALUE", "CEILING.MATH", "FLOOR.MATH", "ISFORMULA", "ISOWEEKNUM",
    "STDEV.S", "STDEV.P", "VAR.S", "VAR.P", "RANK.EQ", "RANK.AVG",
    "PERCENTILE.INC", "PERCENTILE.EXC", "QUARTILE.INC", "QUARTILE.EXC",
    "NORM.S.INV", "NORM.INV", "NORM.S.DIST", "NORM.DIST", "FORECAST.LINEAR",
    "BINOM.DIST", "T.INV", "CHISQ.INV", "COVARIANCE.P", "MODE.SNGL",
    "TEXTBEFORE", "TEXTAFTER", "TEXTSPLIT", "LET", "LAMBDA",
}

_FN_RE = re.compile(r"(?<![A-Za-z0-9_.])([A-Z][A-Z0-9.]*)\s*\(")


def to_ooxml(formula: str) -> str:
    return _FN_RE.sub(
        lambda m: (f"_xlfn.{m.group(1)}(" if m.group(1) in XLFN else m.group(0)),
        formula,
    )


def namespace_workbook(wb) -> None:
    """Apply `to_ooxml` to every formula in the workbook, in place."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = to_ooxml(cell.value)
